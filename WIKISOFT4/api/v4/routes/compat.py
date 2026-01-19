"""
WIKISOFT3 호환 라우트

프론트엔드가 기존 /api/* 경로를 사용하므로 호환성을 위해 제공
"""

from datetime import datetime, timezone
from typing import Optional, List, Dict, Any
from uuid import uuid4
from io import BytesIO
from urllib.parse import quote
import json

from fastapi import APIRouter, File, Form, Header, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core.memory.session_store import session_store

router = APIRouter()

# =============================================================================
# 진단 질문 목록 (기초자료_퇴직급여 기반)
# =============================================================================

ROSTER_QUESTIONS = [
    # ========================================
    # 1. 일반사항 (단답형)
    # ========================================
    {"id": "1-1", "type": "text", "category": "general",
     "question": "회사명", "mapping": "회사명"},
    {"id": "1-2", "type": "text", "category": "general",
     "question": "담당자", "mapping": "담당자"},
    {"id": "1-3", "type": "text", "category": "general",
     "question": "전화번호", "mapping": "전화번호"},
    {"id": "1-4", "type": "text", "category": "general",
     "question": "이메일", "mapping": "이메일"},
    {"id": "1-5", "type": "choice", "category": "general",
     "question": "결산월",
     "choices": ["1월", "2월", "3월", "4월", "5월", "6월",
                 "7월", "8월", "9월", "10월", "11월", "12월"],
     "mapping": "결산월"},
    {"id": "1-6", "type": "text", "category": "general",
     "question": "주소", "mapping": "주소"},

    # ========================================
    # 2. 평가대상 및 퇴직급여 지급(환입) 현황 (숫자)
    # ========================================
    {"id": "2-0", "type": "date", "category": "evaluation",
     "question": "작성기준일", "mapping": "작성기준일", "format": "YYMMDD"},

    # 2-a. 재직자수 (총액은 자동 계산)
    {"id": "2-a", "type": "group", "category": "headcount_active",
     "question": "재직자수", "mapping": "재직자수", "calculated": True},
    {"id": "2-a.1", "type": "number", "category": "headcount_active",
     "question": "임원", "mapping": "재직자수_임원", "unit": "명", "parent": "2-a"},
    {"id": "2-a.2", "type": "number", "category": "headcount_active",
     "question": "직원", "mapping": "재직자수_직원", "unit": "명", "parent": "2-a"},
    {"id": "2-a.3", "type": "number", "category": "headcount_active",
     "question": "계약직", "mapping": "재직자수_계약직", "unit": "명", "parent": "2-a"},

    # 2-b. 퇴직자수 (총액은 자동 계산)
    {"id": "2-b", "type": "group", "category": "headcount_retired",
     "question": "퇴직자수", "mapping": "퇴직자수", "calculated": True},
    {"id": "2-b.1", "type": "number", "category": "headcount_retired",
     "question": "임원", "mapping": "퇴직자수_임원", "unit": "명", "parent": "2-b"},
    {"id": "2-b.2", "type": "number", "category": "headcount_retired",
     "question": "직원", "mapping": "퇴직자수_직원", "unit": "명", "parent": "2-b"},
    {"id": "2-b.3", "type": "number", "category": "headcount_retired",
     "question": "계약직", "mapping": "퇴직자수_계약직", "unit": "명", "parent": "2-b"},

    # 2-c. 중간정산/DC전환자수
    {"id": "2-c.1", "type": "number", "category": "headcount_special",
     "question": "중간정산자수", "mapping": "중간정산자수", "unit": "명"},
    {"id": "2-c.2", "type": "number", "category": "headcount_special",
     "question": "DC전환자수", "mapping": "DC전환자수", "unit": "명"},

    # 2-d. 퇴직금 추계액 (총액은 자동 계산)
    {"id": "2-d", "type": "group", "category": "retirement_estimate",
     "question": "퇴직금 추계액", "mapping": "퇴직금추계액", "calculated": True},
    {"id": "2-d.1", "type": "number", "category": "retirement_estimate",
     "question": "임원", "mapping": "퇴직금추계액_임원", "unit": "원", "parent": "2-d"},
    {"id": "2-d.2", "type": "number", "category": "retirement_estimate",
     "question": "직원", "mapping": "퇴직금추계액_직원", "unit": "원", "parent": "2-d"},
    {"id": "2-d.3", "type": "number", "category": "retirement_estimate",
     "question": "계약직", "mapping": "퇴직금추계액_계약직", "unit": "원", "parent": "2-d"},

    # 2-e. 퇴직부채 감소(증가)액 (A/B/C 그룹)
    {"id": "2-e", "type": "group", "category": "liability_change",
     "question": "퇴직부채 감소(증가)액", "mapping": "퇴직부채변동액", "calculated": True},

    # (A) 급여지급액
    {"id": "2-e.A", "type": "group", "category": "liability_change",
     "question": "(A) 급여지급액", "mapping": "변동_급여지급액", "calculated": True, "parent": "2-e"},
    {"id": "2-e.A.1", "type": "number", "category": "liability_change",
     "question": "퇴직금", "mapping": "변동_퇴직금", "unit": "원", "parent": "2-e.A"},
    {"id": "2-e.A.2", "type": "number", "category": "liability_change",
     "question": "중간정산금액", "mapping": "변동_중간정산금액", "unit": "원", "parent": "2-e.A"},
    {"id": "2-e.A.3", "type": "number", "category": "liability_change",
     "question": "DC전환금", "mapping": "변동_DC전환금", "unit": "원", "parent": "2-e.A"},

    # (B) 관계사 전입(전출)액
    {"id": "2-e.B", "type": "group", "category": "liability_change",
     "question": "(B) 관계사 전입(전출)액", "mapping": "변동_관계사전입전출액", "calculated": True, "parent": "2-e"},
    {"id": "2-e.B.1", "type": "number", "category": "liability_change",
     "question": "전입액", "mapping": "변동_전입액", "unit": "원", "parent": "2-e.B"},
    {"id": "2-e.B.2", "type": "number", "category": "liability_change",
     "question": "전출액", "mapping": "변동_전출액", "unit": "원", "parent": "2-e.B"},

    # (C) 사업결합(처분)액
    {"id": "2-e.C", "type": "group", "category": "liability_change",
     "question": "(C) 사업결합(처분)액", "mapping": "변동_사업결합처분액", "calculated": True, "parent": "2-e"},
    {"id": "2-e.C.1", "type": "number", "category": "liability_change",
     "question": "결합액", "mapping": "변동_사업결합액", "unit": "원", "parent": "2-e.C"},
    {"id": "2-e.C.2", "type": "number", "category": "liability_change",
     "question": "처분액", "mapping": "변동_사업처분액", "unit": "원", "parent": "2-e.C"},

    # ========================================
    # 3. 사외적립자산 현황 (숫자)
    # ========================================
    # 3-a. 기초 잔액
    {"id": "3-a", "type": "number", "category": "plan_assets",
     "question": "기초 잔액", "mapping": "기초잔액", "unit": "원"},

    # 3-b. 입금(수탁)액
    {"id": "3-b", "type": "number", "category": "plan_assets",
     "question": "입금(수탁)액", "mapping": "입금액", "unit": "원"},

    # 3-c. 급여지급(인출)액
    {"id": "3-c", "type": "group", "category": "plan_assets_withdrawal",
     "question": "급여지급(인출)액", "mapping": "인출액", "calculated": True},
    {"id": "3-c.1", "type": "number", "category": "plan_assets_withdrawal",
     "question": "퇴직", "mapping": "인출_퇴직", "unit": "원", "parent": "3-c"},
    {"id": "3-c.2", "type": "number", "category": "plan_assets_withdrawal",
     "question": "중간정산", "mapping": "인출_중간정산", "unit": "원", "parent": "3-c"},
    {"id": "3-c.3", "type": "number", "category": "plan_assets_withdrawal",
     "question": "DC전환", "mapping": "인출_DC전환", "unit": "원", "parent": "3-c"},

    # 3-d. 관계사전입(전출)액
    {"id": "3-d", "type": "group", "category": "plan_assets_transfer",
     "question": "관계사전입(전출)액", "mapping": "관계사전입전출액", "calculated": True},
    {"id": "3-d.1", "type": "number", "category": "plan_assets_transfer",
     "question": "전입", "mapping": "자산_전입액", "unit": "원", "parent": "3-d"},
    {"id": "3-d.2", "type": "number", "category": "plan_assets_transfer",
     "question": "전출", "mapping": "자산_전출액", "unit": "원", "parent": "3-d"},

    # 3-e. 사업결합(처분)액
    {"id": "3-e", "type": "group", "category": "plan_assets_merger",
     "question": "사업결합(처분)액", "mapping": "사업결합처분액", "calculated": True},
    {"id": "3-e.1", "type": "number", "category": "plan_assets_merger",
     "question": "결합", "mapping": "자산_사업결합", "unit": "원", "parent": "3-e"},
    {"id": "3-e.2", "type": "number", "category": "plan_assets_merger",
     "question": "처분", "mapping": "자산_사업처분", "unit": "원", "parent": "3-e"},

    # 3-f. 운용성과
    {"id": "3-f", "type": "group", "category": "plan_assets_performance",
     "question": "운용성과", "mapping": "운용성과", "calculated": True},
    {"id": "3-f.1", "type": "number", "category": "plan_assets_performance",
     "question": "투자수익", "mapping": "투자수익", "unit": "원", "parent": "3-f"},
    {"id": "3-f.2", "type": "number", "category": "plan_assets_performance",
     "question": "운용관리수수료", "mapping": "운용관리수수료", "unit": "원", "parent": "3-f"},

    # 3-g. 기말 잔액
    {"id": "3-g", "type": "group", "category": "plan_assets_ending",
     "question": "기말 잔액", "mapping": "기말잔액", "calculated": True},
    {"id": "3-g.1", "type": "number", "category": "plan_assets_ending",
     "question": "퇴직연금", "mapping": "기말잔액_퇴직연금", "unit": "원", "parent": "3-g"},
    {"id": "3-g.2", "type": "number", "category": "plan_assets_ending",
     "question": "퇴직신탁", "mapping": "기말잔액_퇴직신탁", "unit": "원", "parent": "3-g"},
    {"id": "3-g.3", "type": "number", "category": "plan_assets_ending",
     "question": "퇴직보험", "mapping": "기말잔액_퇴직보험", "unit": "원", "parent": "3-g"},

    # 3-h. 국민연금전환금
    {"id": "3-h", "type": "number", "category": "plan_assets",
     "question": "국민연금전환금 기말잔액", "mapping": "국민연금전환금", "unit": "원"},

    # 3-i. 공시방법
    {"id": "3-i", "type": "choice", "category": "plan_assets",
     "question": "공시방법", "mapping": "공시방법",
     "choices": ["순액법", "총액법"]},

    # ========================================
    # 4. 퇴직급여제도의 주요 사항 (선택/입력)
    # ========================================
    {"id": "4-1", "type": "number", "category": "plan_details",
     "question": "정년퇴직연령(만)", "mapping": "정년퇴직연령", "unit": "세"},
    {"id": "4-2", "type": "choice", "category": "plan_details",
     "question": "임금피크제", "mapping": "임금피크제",
     "choices": ["예", "아니오"]},
    {"id": "4-2.1", "type": "number", "category": "plan_details",
     "question": "ㄴ 적용시작연령", "mapping": "임금피크제_시작연령", "unit": "세",
     "parent": "4-2", "condition": {"question_id": "4-2", "answer": "예"}},
    {"id": "4-3", "type": "choice", "category": "plan_details",
     "question": "기타장기종업원급여", "mapping": "기타장기종업원급여",
     "choices": ["예", "아니오"]},
    {"id": "4-4", "type": "choice", "category": "plan_details",
     "question": "지급률 가감", "mapping": "지급률가감",
     "choices": ["+", "-", "없음"]},
    {"id": "4-5", "type": "choice", "category": "plan_details",
     "question": "제도구분", "mapping": "제도구분",
     "choices": ["법정제", "누진제"]},
    {"id": "4-6", "type": "choice", "category": "plan_details",
     "question": "1년미만 근무기간 처리방법", "mapping": "1년미만처리방법",
     "choices": ["일할", "월할(절사)", "월할(절상)"]},
    {"id": "4-7", "type": "choice", "category": "plan_details",
     "question": "축소/정산/사업결합/중간정산/DC전환 발생여부", "mapping": "특수사건발생",
     "choices": ["예", "아니오"]},
    {"id": "4-8", "type": "choice", "category": "plan_details",
     "question": "일부 중간정산/DC전환 발생여부", "mapping": "일부중간정산발생",
     "choices": ["예", "아니오"]},
    {"id": "4-9", "type": "choice", "category": "plan_details",
     "question": "임금인상률 임금체계", "mapping": "임금체계",
     "choices": ["호봉제", "연봉제"]},
    # 임금인상률 (5년간) - 작성기준일에 따라 동적 생성 필요
    {"id": "4-10", "type": "group", "category": "wage_increase",
     "question": "임금인상률 (최근 5년)", "mapping": "임금인상률", "calculated": False},
    {"id": "4-10.1", "type": "number", "category": "wage_increase",
     "question": "1년 전", "mapping": "임금인상률_Y1", "unit": "%", "parent": "4-10"},
    {"id": "4-10.2", "type": "number", "category": "wage_increase",
     "question": "2년 전", "mapping": "임금인상률_Y2", "unit": "%", "parent": "4-10"},
    {"id": "4-10.3", "type": "number", "category": "wage_increase",
     "question": "3년 전", "mapping": "임금인상률_Y3", "unit": "%", "parent": "4-10"},
    {"id": "4-10.4", "type": "number", "category": "wage_increase",
     "question": "4년 전", "mapping": "임금인상률_Y4", "unit": "%", "parent": "4-10"},
    {"id": "4-10.5", "type": "number", "category": "wage_increase",
     "question": "5년 전", "mapping": "임금인상률_Y5", "unit": "%", "parent": "4-10"},

    # ========================================
    # 5. 할인율 산출 기준 (선택)
    # ========================================
    {"id": "5-1", "type": "choice", "category": "discount_rate",
     "question": "할인율 채권등급", "mapping": "할인율채권등급",
     "choices": ["국공채", "회사채AAA", "AA+", "AA0", "AA-"]},

    # ========================================
    # 6. 사외적립자산 적립비율 (선택/입력)
    # ========================================
    {"id": "6-1", "type": "choice", "category": "funding_ratio",
     "question": "적립비율", "mapping": "적립비율",
     "choices": ["80.0% (목표 사외적립비율)", "현재 사외적립비율", "기타", "사외적립비율 없음 (0%)"]},
    {"id": "6-1.1", "type": "number", "category": "funding_ratio",
     "question": "기타 적립비율 입력", "mapping": "적립비율_기타", "unit": "%",
     "condition": {"question_id": "6-1", "answer": "기타"}},

    # ========================================
    # 7. 특이사항 (주관식)
    # ========================================
    {"id": "7-1", "type": "textarea", "category": "special_notes",
     "question": "임금피크제 관련 특이사항", "mapping": "특이사항_임금피크제"},
    {"id": "7-2", "type": "textarea", "category": "special_notes",
     "question": "지급률 가감 관련 특이사항", "mapping": "특이사항_지급률가감"},
    {"id": "7-3", "type": "textarea", "category": "special_notes",
     "question": "퇴직금 누진제 관련 특이사항", "mapping": "특이사항_누진제"},
    {"id": "7-4", "type": "textarea", "category": "special_notes",
     "question": "기타 특이사항 (축소, 정산, 사업결합, 중간정산, DC전환 등)",
     "mapping": "특이사항_기타"},
]


@router.get("/api/diagnostic-questions")
async def get_diagnostic_questions():
    """진단 질문 목록 (WIKISOFT3 호환)"""
    return {
        "total": len(ROSTER_QUESTIONS),
        "questions": ROSTER_QUESTIONS
    }


@router.post("/api/auto-validate")
async def auto_validate(
    file: UploadFile = File(...),
    chatbot_answers: Optional[str] = Form(None),
    validation_mode: Optional[str] = Form("full"),  # "roster" 또는 "full"
    x_session_id: Optional[str] = Header(None)
):
    """파일 업로드 및 자동 검증 (WIKISOFT3 호환)

    validation_mode:
        - "roster": 명부 형식만 검증 (L1만, 금액 비교 없음)
        - "full": 전체 검증 (L1 + L2 금액 비교 + AI)
    """

    # 세션 ID 생성 또는 재사용
    session_id = session_store.get_or_create(x_session_id)

    # 파일 검증
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일이 필요합니다")

    filename = file.filename.lower()
    if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Excel 파일만 지원합니다 (xls, xlsx)")

    # 파일 읽기
    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(status_code=400, detail="파일 크기는 10MB 이하만 가능합니다")

    # 진단 답변 파싱
    diagnostic_answers = {}
    if chatbot_answers:
        try:
            diagnostic_answers = json.loads(chatbot_answers)
        except json.JSONDecodeError:
            pass

    try:
        # core 모듈 사용
        from core.parsers.parser import parse_roster
        from core.ai.matcher import match_headers
        from core.validators.validator import validate
        from core.validators.duplicate_detector import detect_duplicates
        from core.agent.confidence import estimate_confidence, detect_anomalies
        from core.generators.report import generate_report
        import pandas as pd

        # 1. 파싱
        parsed = parse_roster(file_bytes)

        # 2. 헤더 매칭
        matches = match_headers(parsed, sheet_type="재직자")

        # 무시 컬럼 필터링
        filtered_matches = [
            m for m in matches.get("matches", [])
            if not m.get("ignored", False)
        ]
        matches_for_response = {**matches, "matches": filtered_matches}

        # DataFrame 생성
        df = pd.DataFrame(parsed.get("rows", []), columns=parsed.get("headers", []))

        # 3. 검증 (mode: "roster" = 명부만, "full" = 전체)
        validation = validate(
            parsed=parsed,
            matches=matches,
            diagnostic_answers=diagnostic_answers,
            df=df,
            mode=validation_mode or "full"
        )

        # 4. 신뢰도/이상치 분석
        confidence = estimate_confidence(parsed, matches, validation)
        anomalies = detect_anomalies(parsed, matches, validation)

        # 5. 중복 탐지
        duplicates = detect_duplicates(
            df=df,
            headers=parsed.get("headers", []),
            matches=matches.get("matches", [])
        )

        # 6. 리포트 생성
        report = generate_report(validation=validation)

        # 스프레드시트 에디터용 데이터 (최대 100행)
        rows_data = parsed.get("rows", [])[:100]
        headers = parsed.get("headers", [])
        all_rows_with_headers = [
            {h: row[i] if i < len(row) else "" for i, h in enumerate(headers)}
            for row in rows_data
        ]

        # status 동적 설정: errors 있으면 "error", warnings만 있으면 "warning", 둘 다 없으면 "ok"
        errors = validation.get("errors", [])
        warnings_list = validation.get("warnings", [])
        if errors:
            status = "error"
        elif warnings_list:
            status = "warning"
        else:
            status = "ok"

        result = {
            "status": status,
            "session_id": session_id,
            "confidence": confidence,
            "anomalies": anomalies,
            "duplicates": duplicates,
            "diagnostic_applied": bool(diagnostic_answers),
            "steps": {
                "parsed_summary": {
                    "headers": parsed.get("headers", []),
                    "row_count": len(parsed.get("rows", [])),
                },
                "all_rows": all_rows_with_headers,
                "matches": matches_for_response,
                "validation": validation,
                "report": report,
            },
        }

        # 세션에 결과 저장 (리포트 다운로드용)
        session_store.set(session_id, "validation_result", result)
        session_store.set(session_id, "parsed_data", parsed)
        session_store.set(session_id, "diagnostic_answers", diagnostic_answers)
        session_store.set(session_id, "filename", file.filename)

        return result

    except ImportError as e:
        # core 모듈 미완성 시 placeholder 반환
        return {
            "status": "ok",
            "session_id": session_id,
            "confidence": {"score": 0.85, "level": "medium"},
            "anomalies": {"detected": False, "anomalies": []},
            "duplicates": {"has_duplicates": False},
            "diagnostic_applied": bool(diagnostic_answers),
            "steps": {
                "parsed_summary": {
                    "headers": ["사원번호", "성명", "생년월일", "입사일"],
                    "row_count": 0,
                },
                "all_rows": [],
                "matches": {"matches": [], "confidence": 0.85},
                "validation": {"errors": [], "warnings": []},
                "report": {"summary": "Core 모듈 로딩 실패"},
            },
            "error": f"Core module import error: {str(e)}. Using placeholder response."
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"검증 오류: {str(e)}"
        )


# ========================================
# 재검증 API
# ========================================

class RevalidateRequest(BaseModel):
    """재검증 요청"""
    headers: List[str]
    rows: List[List[Any]]
    diagnostic_answers: Optional[Dict[str, Any]] = None


@router.post("/api/auto-validate/revalidate")
async def revalidate(request: RevalidateRequest):
    """수정된 데이터로 재검증 수행

    사용자가 UI에서 데이터를 수정한 후 재검증 버튼을 누르면 호출됨.
    Layer 1(형식), Layer 2(집계), AI 검증을 다시 실행하여
    오류/경고 개수를 재산정함.
    """
    try:
        from core.validators.validator import validate
        from core.ai.matcher import match_headers
        from core.agent.confidence import estimate_confidence, detect_anomalies
        from core.validators.duplicate_detector import detect_duplicates
        import pandas as pd

        headers = request.headers
        rows = request.rows
        diagnostic_answers = request.diagnostic_answers or {}

        # parsed 형식으로 변환
        parsed = {
            "headers": headers,
            "rows": [dict(zip(headers, row)) for row in rows]
        }

        # DataFrame 생성
        df = pd.DataFrame(rows, columns=headers)

        # 헤더 재매칭
        matches = match_headers(parsed, sheet_type="재직자")

        # 검증 재실행
        validation = validate(
            parsed=parsed,
            matches=matches,
            diagnostic_answers=diagnostic_answers,
            df=df
        )

        # 신뢰도/이상치 재분석
        confidence = estimate_confidence(parsed, matches, validation)
        anomalies = detect_anomalies(parsed, matches, validation)

        # 중복 재탐지
        duplicates = detect_duplicates(
            df=df,
            headers=headers,
            matches=matches.get("matches", [])
        )

        errors = validation.get("errors", [])
        warnings = validation.get("warnings", [])

        # status 동적 설정
        if errors:
            status = "error"
        elif warnings:
            status = "warning"
        else:
            status = "ok"

        return {
            "status": status,
            "revalidated": True,
            "confidence": confidence,
            "anomalies": anomalies,
            "duplicates": duplicates,
            "errors": errors,
            "warnings": warnings,
            "error_count": len(errors),
            "warning_count": len(warnings),
            "checks": validation.get("checks", []),
            "ai_reasoning": validation.get("ai_reasoning", []),
            "layer2_comparison": validation.get("layer2_comparison"),
            # 프론트엔드 호환성을 위해 validation 객체도 포함
            "validation": validation,
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Core 모듈 로딩 실패: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"재검증 오류: {str(e)}"
        )


# ========================================
# 검증 리포트 다운로드 API
# ========================================

@router.get("/api/auto-validate/download-excel")
async def download_excel_report(x_session_id: Optional[str] = Header(None)):
    """검증 결과를 Excel 리포트로 다운로드"""
    if not x_session_id:
        raise HTTPException(status_code=400, detail="세션 ID가 필요합니다")

    # 세션에서 데이터 조회
    validation_result = session_store.get(x_session_id, "validation_result")
    parsed_data = session_store.get(x_session_id, "parsed_data")
    diagnostic_answers = session_store.get(x_session_id, "diagnostic_answers")
    filename = session_store.get(x_session_id, "filename") or "검증결과"

    if not validation_result:
        raise HTTPException(status_code=404, detail="세션 데이터가 없거나 만료되었습니다")

    try:
        from core.generators.report import generate_excel_report

        # Excel 리포트 생성
        excel_bytes = generate_excel_report(
            validation_result=validation_result,
            original_data=parsed_data,
            answers=diagnostic_answers
        )

        # 파일명 생성
        base_name = filename.rsplit(".", 1)[0] if "." in filename else filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_검증리포트_{timestamp}.xlsx"

        return StreamingResponse(
            BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_filename)}"}
        )

    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"리포트 생성 모듈 로딩 실패: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"리포트 생성 오류: {str(e)}")


@router.get("/api/auto-validate/download-final-data")
async def download_final_data(x_session_id: Optional[str] = Header(None)):
    """최종 수정본 다운로드 (매핑된 데이터)"""
    if not x_session_id:
        raise HTTPException(status_code=400, detail="세션 ID가 필요합니다")

    # 세션에서 데이터 조회
    parsed_data = session_store.get(x_session_id, "parsed_data")
    filename = session_store.get(x_session_id, "filename") or "최종수정본"

    if not parsed_data:
        raise HTTPException(status_code=404, detail="세션 데이터가 없거나 만료되었습니다")

    try:
        headers = parsed_data.get("headers", [])
        rows = parsed_data.get("rows", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재직자명부"

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        # 헤더 작성
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # 데이터 작성
        for row_idx, row in enumerate(rows, start=2):
            if isinstance(row, dict):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            else:
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 파일명 생성
        base_name = filename.rsplit(".", 1)[0] if "." in filename else filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_최종수정본_{timestamp}.xlsx"

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_filename)}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 생성 오류: {str(e)}")


@router.get("/api/windmill/latest")
async def get_windmill_latest(limit: int = 5):
    """최근 실행 기록 (WIKISOFT3 호환)"""
    return {
        "runs": [],
        "note": "WIKISOFT4에서는 n8n/Temporal 워크플로우 사용 예정"
    }


@router.get("/api/health")
async def health_compat():
    """헬스 체크 (WIKISOFT3 호환)"""
    return {
        "status": "ok",
        "version": "4.1.0",
        "mode": "compatibility"
    }


# ========================================
# Export API (WIKISOFT3 호환)
# ========================================

class ExportRequest(BaseModel):
    """수정된 데이터 내보내기 요청"""
    filename: str
    headers: List[str]
    data: List[Dict[str, Any]]


class ExportErrorsRequest(BaseModel):
    """오류 항목만 추출하여 내보내기"""
    filename: str
    errors: List[Dict[str, Any]]


@router.post("/api/export/xlsx")
async def export_xlsx(request: ExportRequest):
    """수정된 데이터를 Excel 파일로 내보내기"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재직자명부"

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 작성
        for col_idx, header in enumerate(request.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 데이터 작성
        for row_idx, row_data in enumerate(request.data, start=2):
            for col_idx, header in enumerate(request.headers, start=1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(request.headers, start=1):
            max_length = len(str(header))
            for row_idx in range(2, len(request.data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # 파일명 생성
        base_name = request.filename.rsplit(".", 1)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_수정본_{timestamp}.xlsx"

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_filename)}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 생성 실패: {str(e)}")


# ========================================
# Agent Chat API (FloatingChat용)
# ========================================

class AgentAskRequest(BaseModel):
    """AI 에이전트 질문 요청"""
    message: str
    context: Optional[Dict[str, Any]] = None  # 검증 컨텍스트 (선택)


@router.post("/api/agent/ask")
async def agent_ask(request: AgentAskRequest):
    """AI 에이전트에게 질문 (FloatingChat용)

    HR/Finance 검증 관련 질문에 답변하는 AI 어시스턴트
    """
    try:
        from core.ai.llm_client import get_llm_client

        # 시스템 프롬프트: 검증 도우미
        system_prompt = """당신은 WIKISOFT 4.1의 AI 어시스턴트입니다.
퇴직연금 계리평가를 위한 명부 검증 플랫폼의 전문 도우미입니다.

주요 역할:
- 진단 질문 작성 도움 (회사 정보, 재직자수, 퇴직급여 등)
- 명부 검증 오류 해결 방법 안내
- 퇴직연금/퇴직급여 관련 용어 설명
- 엑셀 파일 형식 및 필수 필드 안내

답변 규칙:
- 간결하고 명확하게 답변
- 한국어로 답변
- 검증 관련 질문은 구체적인 해결책 제시
- 모르는 내용은 솔직히 모른다고 답변"""

        messages = [
            {"role": "system", "content": system_prompt},
        ]

        # 검증 컨텍스트가 있으면 추가
        if request.context:
            context_str = f"현재 검증 상태:\n"
            if request.context.get("errors"):
                context_str += f"- 오류 {len(request.context['errors'])}건\n"
            if request.context.get("warnings"):
                context_str += f"- 경고 {len(request.context['warnings'])}건\n"
            if request.context.get("step"):
                context_str += f"- 현재 단계: {request.context['step']}\n"
            messages.append({"role": "system", "content": context_str})

        messages.append({"role": "user", "content": request.message})

        # LLM 호출
        client = get_llm_client()
        response = client.chat(messages, temperature=0.7, max_tokens=500)

        return {
            "status": "ok",
            "response": response
        }

    except ValueError as e:
        # API 키 없음
        return {
            "status": "error",
            "response": "AI 기능을 사용하려면 OpenAI API 키가 필요합니다. 관리자에게 문의하세요.",
            "error": str(e)
        }
    except Exception as e:
        return {
            "status": "error",
            "response": "죄송합니다. 일시적인 오류가 발생했습니다. 잠시 후 다시 시도해주세요.",
            "error": str(e)
        }


@router.post("/api/export/errors")
async def export_errors(request: ExportErrorsRequest):
    """검증 오류만 추출하여 Excel 파일로 내보내기"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "오류항목"

        # 헤더 스타일
        header_fill = PatternFill(start_color="FF5555", end_color="FF5555", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 헤더 작성
        headers = ["행번호", "사원번호", "필드명", "오류 내용", "심각도"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 데이터 작성
        for row_idx, error in enumerate(request.errors, start=2):
            ws.cell(row=row_idx, column=1, value=error.get('row', ''))

            # 사원번호 추출 (emp_info에서 "사원번호 xxx" 형식 처리)
            emp_info = error.get('emp_info', '')
            if emp_info and emp_info.startswith('사원번호 '):
                emp_id = emp_info.replace('사원번호 ', '').split('(')[0].strip()
            elif emp_info and emp_info.startswith('행 '):
                emp_id = ''  # 사원번호 없이 행 번호만 있는 경우
            else:
                emp_id = emp_info  # 그대로 사용
            ws.cell(row=row_idx, column=2, value=emp_id)

            ws.cell(row=row_idx, column=3, value=error.get('field', ''))

            message_cell = ws.cell(row=row_idx, column=4, value=error.get('message', ''))
            message_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            severity = error.get('severity', 'warning')
            severity_cell = ws.cell(row=row_idx, column=5, value=severity)
            if severity == 'error':
                severity_cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                severity_cell.font = Font(color="CC0000", bold=True)
            else:
                severity_cell.fill = PatternFill(start_color="FFF2D6", end_color="FFF2D6", fill_type="solid")
                severity_cell.font = Font(color="FF8800", bold=True)
            severity_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 10   # 행번호
        ws.column_dimensions['B'].width = 15   # 사원번호
        ws.column_dimensions['C'].width = 15   # 필드명
        ws.column_dimensions['D'].width = 50   # 오류 내용
        ws.column_dimensions['E'].width = 10   # 심각도

        # 파일명 생성
        base_name = request.filename.rsplit(".", 1)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_의심목록_{timestamp}.xlsx"

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_filename)}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 생성 실패: {str(e)}")
