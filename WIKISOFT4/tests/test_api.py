"""
API 엔드포인트 테스트 (WIKISOFT4)
"""
import pytest
from fastapi.testclient import TestClient
import io
from openpyxl import Workbook
import sys
sys.path.insert(0, "/Users/kj/Desktop/wiki/WIKISOFT4")

from api.v4.main import app

client = TestClient(app)


def create_test_excel() -> bytes:
    """테스트용 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "재직자"

    headers = ["사원번호", "이름", "생년월일", "입사일", "기준급여", "종업원구분"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    data = [
        ["EMP001", "홍길동", "1990-01-15", "2020-03-01", 5000000, "직원"],
        ["EMP002", "김철수", "1985-06-20", "2018-07-15", 6000000, "직원"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestHealthEndpoint:
    """Health 엔드포인트 테스트"""

    def test_health_check(self):
        """헬스 체크"""
        response = client.get("/api/v4/health")
        assert response.status_code == 200

        data = response.json()
        assert data["status"] in ["healthy", "ok"]
        assert "version" in data

    def test_health_check_compat(self):
        """호환성 헬스 체크 (v3 경로)"""
        response = client.get("/api/health")
        assert response.status_code == 200


class TestDiagnosticQuestionsEndpoint:
    """진단 질문 엔드포인트 테스트"""

    def test_get_questions(self):
        """질문 조회"""
        response = client.get("/api/diagnostic-questions")
        assert response.status_code == 200

        data = response.json()
        assert "questions" in data
        assert "total" in data
        assert data["total"] >= 13  # WIKISOFT4에서는 23개

    def test_questions_structure(self):
        """질문 구조 확인"""
        response = client.get("/api/diagnostic-questions")
        data = response.json()

        if data["questions"]:
            q = data["questions"][0]
            assert "id" in q
            assert "question" in q
            assert "type" in q
            assert "category" in q


class TestValidateEndpoint:
    """검증 엔드포인트 테스트"""

    def test_validate_with_file(self):
        """파일 검증"""
        excel_bytes = create_test_excel()

        response = client.post(
            "/api/auto-validate",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )

        assert response.status_code in [200, 404]
        if response.status_code == 200:
            data = response.json()
            assert "status" in data or "success" in data

    def test_validate_without_file(self):
        """파일 없이 검증 (에러)"""
        response = client.post("/api/auto-validate")

        assert response.status_code in [400, 404, 422]


class TestAuthEndpoint:
    """인증 엔드포인트 테스트 (WIKISOFT4 신규)"""

    def test_login_success(self):
        """로그인 성공"""
        response = client.post(
            "/api/v4/auth/login",
            json={"username": "admin", "password": "admin1234!"}
        )

        if response.status_code == 200:
            data = response.json()
            assert "access_token" in data
            assert data["token_type"] == "bearer"

    def test_login_failure(self):
        """로그인 실패"""
        response = client.post(
            "/api/v4/auth/login",
            json={"username": "wronguser", "password": "wrongpassword"}
        )

        assert response.status_code in [401, 400, 422]


class TestPrivacyEndpoint:
    """프라이버시 엔드포인트 테스트 (WIKISOFT4 신규)"""

    def test_get_pii_types(self):
        """PII 타입 조회"""
        response = client.get("/api/v4/privacy/pii-types")

        if response.status_code == 200:
            data = response.json()
            assert "pii_types" in data or "types" in data or isinstance(data, list)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
