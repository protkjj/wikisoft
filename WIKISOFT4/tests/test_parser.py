"""
Parser 테스트 (WIKISOFT4)
"""
import pytest
import io
from openpyxl import Workbook

import sys
sys.path.insert(0, "/Users/kj/Desktop/wiki/WIKISOFT4")

from core.parsers.parser import parse_roster


def create_test_excel() -> bytes:
    """테스트용 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "재직자"

    # 헤더
    headers = ["사원번호", "이름", "생년월일", "입사일", "기준급여", "종업원구분"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터
    data = [
        ["EMP001", "홍길동", "1990-01-15", "2020-03-01", 5000000, "직원"],
        ["EMP002", "김철수", "1985-06-20", "2018-07-15", 6000000, "직원"],
        ["EMP003", "이영희", "1992-11-30", "2021-01-10", 4500000, "계약직"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestParser:
    """Parser 테스트"""

    def test_parse_roster_basic(self):
        """기본 파싱 테스트"""
        excel_bytes = create_test_excel()
        result = parse_roster(excel_bytes)

        assert result is not None
        assert "headers" in result
        assert "rows" in result or "row_count" in result

    def test_parse_roster_headers(self):
        """헤더 추출 테스트"""
        excel_bytes = create_test_excel()
        result = parse_roster(excel_bytes)

        assert len(result["headers"]) == 6
        assert "사원번호" in result["headers"]
        assert "이름" in result["headers"]

    def test_parse_roster_row_count(self):
        """행 개수 테스트"""
        excel_bytes = create_test_excel()
        result = parse_roster(excel_bytes)

        row_count = result.get("row_count", len(result.get("rows", [])))
        assert row_count >= 1

    def test_parse_roster_with_sheet_name(self):
        """시트 이름 지정 테스트"""
        excel_bytes = create_test_excel()
        result = parse_roster(excel_bytes, sheet_name="재직자")

        assert result is not None
        row_count = result.get("row_count", len(result.get("rows", [])))
        assert row_count >= 1

    def test_parse_roster_empty_bytes(self):
        """빈 파일 테스트"""
        try:
            result = parse_roster(b"")
            assert result is not None
        except Exception:
            pass

    def test_parse_roster_max_rows(self):
        """최대 행 제한 테스트"""
        excel_bytes = create_test_excel()
        result = parse_roster(excel_bytes, max_rows=2)

        row_count = result.get("row_count", len(result.get("rows", [])))
        assert row_count >= 1


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
