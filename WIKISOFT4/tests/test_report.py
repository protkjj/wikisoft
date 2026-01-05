"""
리포트 생성 테스트 (검증 리포트)
"""

import pytest
from io import BytesIO
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from core.generators import generate_excel_report


class TestValidationReport:
    """검증 리포트 테스트"""

    @pytest.fixture
    def sample_validation_result(self):
        """샘플 검증 결과"""
        return {
            "status": "ok",
            "confidence": {"score": 0.92, "grade": "A"},
            "steps": {
                "parsed_summary": {"row_count": 10},
                "matches": {
                    "matches": [
                        {"source": "사원번호", "target": "employee_id", "confidence": 0.95},
                        {"source": "이름", "target": "name", "confidence": 0.98},
                    ]
                },
            },
            "anomalies": {
                "detected": False,
                "anomalies": [],
            },
        }

    @pytest.fixture
    def sample_original_data(self):
        """샘플 원본 데이터"""
        return {
            "headers": ["사원번호", "이름", "생년월일", "입사일", "기준급여"],
            "rows": [
                ["E001", "홍길동", "1985-03-15", "2015-06-01", 5000000],
                ["E002", "김영희", "1990-07-22", "2018-03-01", 4500000],
            ],
        }

    def test_generate_validation_report(self, sample_validation_result, sample_original_data):
        """검증 리포트 생성 테스트"""
        excel_bytes = generate_excel_report(
            sample_validation_result,
            sample_original_data
        )

        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Excel 파일 유효성 확인
        wb = load_workbook(BytesIO(excel_bytes))
        assert "검증 요약" in wb.sheetnames
        assert "헤더 매칭" in wb.sheetnames

    def test_report_with_anomalies(self, sample_original_data):
        """이상 탐지 포함 리포트"""
        result = {
            "status": "warning",
            "confidence": {"score": 0.75, "grade": "B"},
            "steps": {
                "parsed_summary": {"row_count": 2},
                "matches": {"matches": []},
            },
            "anomalies": {
                "detected": True,
                "anomalies": [
                    {"type": "outlier", "severity": "high", "message": "급여 이상치", "field": "기준급여"},
                ],
                "recommendation": "검토 필요",
            },
        }

        excel_bytes = generate_excel_report(result, sample_original_data)
        wb = load_workbook(BytesIO(excel_bytes))
        assert "이상 탐지" in wb.sheetnames


class TestReportAPI:
    """리포트 API 테스트"""

    @pytest.fixture
    def client(self):
        from api.v4.main import app
        return TestClient(app)

    def test_validation_report_endpoint(self, client):
        """검증 리포트 API 테스트"""
        payload = {
            "validation_result": {
                "status": "ok",
                "confidence": {"score": 0.9, "grade": "A"},
                "steps": {"parsed_summary": {"row_count": 5}, "matches": {"matches": []}},
                "anomalies": {"detected": False, "anomalies": []},
            },
            "filename": "test"
        }
        response = client.post("/api/v4/report/validation", json=payload)
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
