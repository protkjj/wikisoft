"""
Excel 리포트 생성기

검증 결과를 기반으로 Excel 파일 생성:
- 빨간색: 오류 (필수값 누락, 데이터 불일치)
- 노란색: 경고 (5% 이상 차이, 이상치 탐지)
- 초록색: 정상
"""

from typing import Any, Dict, List, Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json


# 셀 스타일 정의
STYLE_ERROR = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 빨간색
STYLE_WARNING = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")  # 노란색
STYLE_SUCCESS = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")  # 초록색
STYLE_HEADER = PatternFill(start_color="4C6EF5", end_color="4C6EF5", fill_type="solid")  # 파란색

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(size=10)
FONT_ERROR = Font(color="8B0000", bold=True)
FONT_WARNING = Font(color="8B4513")

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')


def generate_report(validation: Dict[str, Any]) -> Dict[str, Any]:
    """JSON 리포트 생성"""
    return {
        "report_type": "json",
        "summary": {
            "passed": validation.get("passed"),
            "warnings": validation.get("warnings", []),
            "checks": validation.get("checks", []),
        },
    }


def generate_excel_report(
    validation_result: Dict[str, Any],
    original_data: Optional[Dict[str, Any]] = None,
    answers: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    검증 결과를 기반으로 Excel 리포트 생성
    
    Args:
        validation_result: 검증 결과 (steps, confidence, anomalies 등)
        original_data: 원본 파싱 데이터 (headers, rows)
        answers: 사용자 진단 답변
    
    Returns:
        Excel 파일 bytes
    """
    wb = Workbook()
    
    # 1. 요약 시트
    ws_summary = wb.active
    ws_summary.title = "검증 요약"
    _create_summary_sheet(ws_summary, validation_result, answers)
    
    # 2. 매칭 결과 시트
    ws_matching = wb.create_sheet("헤더 매칭")
    _create_matching_sheet(ws_matching, validation_result)

    # 3. 인원/금액 비교 시트 (Layer 2)
    layer2_data = validation_result.get("layer2_comparison")
    if layer2_data and layer2_data.get("total_checks", 0) > 0:
        ws_layer2 = wb.create_sheet("인원금액 비교")
        _create_layer2_sheet(ws_layer2, layer2_data)

    # 4. 이상 탐지 시트
    if validation_result.get("anomalies", {}).get("detected"):
        ws_anomalies = wb.create_sheet("이상 탐지")
        _create_anomalies_sheet(ws_anomalies, validation_result)
    
    # 4. 원본 데이터 시트 (하이라이팅 포함)
    if original_data and original_data.get("rows"):
        ws_data = wb.create_sheet("검증된 데이터")
        _create_data_sheet(ws_data, original_data, validation_result)
    
    # 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_summary_sheet(ws, validation_result: Dict, answers: Optional[Dict]):
    """요약 시트 생성"""
    # 제목
    ws.merge_cells('A1:D1')
    ws['A1'] = "🏢 OneCheck 검증 결과 리포트"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = ALIGN_CENTER
    
    # 검증 상태
    row = 3
    status = validation_result.get("status", "unknown")
    confidence = validation_result.get("confidence", {})

    # 에러/경고 개수 계산
    validation_data = validation_result.get("steps", {}).get("validation", {})
    error_count = len(validation_data.get("errors", []))
    warning_count = len([w for w in validation_data.get("warnings", []) if isinstance(w, dict)])

    headers = ["항목", "값", "상태", "설명"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    # status 이모지 결정
    if status == "ok":
        status_emoji = "✅"
        status_desc = "검증 통과"
    elif status == "warning":
        status_emoji = "⚠️"
        status_desc = "경고 있음"
    else:
        status_emoji = "❌"
        status_desc = "오류 있음"

    # 데이터 행
    summary_data = [
        ("검증 상태", status, status_emoji, status_desc),
        ("오류 개수", error_count, "❌" if error_count > 0 else "✅", "필수값 누락, 형식 오류 등"),
        ("경고 개수", warning_count, "⚠️" if warning_count > 0 else "✅", "최저임금 미달, 중복 등"),
        ("신뢰도 점수", f"{confidence.get('score', 0) * 100:.1f}%",
         "✅" if confidence.get('score', 0) >= 0.8 else "⚠️", confidence.get('grade', '')),
        ("분석 행 수", validation_result.get("steps", {}).get("parsed_summary", {}).get("row_count", 0),
         "✅", "파싱된 데이터 행 수"),
        ("이상 탐지", len(validation_result.get("anomalies", {}).get("anomalies", [])),
         "⚠️" if validation_result.get("anomalies", {}).get("detected") else "✅",
         validation_result.get("anomalies", {}).get("recommendation", "")),
    ]
    
    for data_row in summary_data:
        row += 1
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_LEFT
            
            # 상태 컬럼 스타일링
            if col == 3:
                if "❌" in str(value):
                    cell.fill = STYLE_ERROR
                elif "⚠️" in str(value):
                    cell.fill = STYLE_WARNING
                else:
                    cell.fill = STYLE_SUCCESS
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    
    # 진단 답변 섹션
    if answers:
        row += 3
        ws.cell(row=row, column=1, value="📋 진단 답변").font = Font(bold=True, size=12)
        row += 1

        for q_id, answer in answers.items():
            row += 1
            ws.cell(row=row, column=1, value=q_id)
            ws.cell(row=row, column=2, value=str(answer))

    # 검증 이슈 목록 섹션 (errors + warnings)
    validation = validation_result.get("steps", {}).get("validation", {})
    errors = validation.get("errors", [])
    warnings_list = validation.get("warnings", [])

    if errors or warnings_list:
        row += 3
        ws.cell(row=row, column=1, value="🚨 검증 이슈 목록").font = Font(bold=True, size=12)
        row += 1

        # 이슈 테이블 헤더
        issue_headers = ["구분", "행번호", "사원번호", "필드", "내용"]
        for col, header in enumerate(issue_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = STYLE_HEADER
            cell.font = FONT_HEADER
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

        # 오류 목록
        for error in errors:
            row += 1
            ws.cell(row=row, column=1, value="❌ 오류").fill = STYLE_ERROR
            ws.cell(row=row, column=1).border = BORDER_THIN

            row_num = error.get("row", "")
            ws.cell(row=row, column=2, value=row_num if row_num else "-").border = BORDER_THIN

            emp_info = error.get("emp_info", "")
            if emp_info and emp_info.startswith("사원번호 "):
                emp_id = emp_info.replace("사원번호 ", "").split("(")[0].strip()
            else:
                emp_id = emp_info or "-"
            ws.cell(row=row, column=3, value=emp_id).border = BORDER_THIN

            ws.cell(row=row, column=4, value=error.get("field", error.get("column", ""))).border = BORDER_THIN
            ws.cell(row=row, column=5, value=error.get("message", error.get("error", ""))).border = BORDER_THIN

        # 경고 목록
        for warning in warnings_list:
            if isinstance(warning, dict):
                row += 1
                ws.cell(row=row, column=1, value="⚠️ 경고").fill = STYLE_WARNING
                ws.cell(row=row, column=1).border = BORDER_THIN

                row_num = warning.get("row", "")
                ws.cell(row=row, column=2, value=row_num if row_num else "-").border = BORDER_THIN

                emp_info = warning.get("emp_info", "")
                if emp_info and emp_info.startswith("사원번호 "):
                    emp_id = emp_info.replace("사원번호 ", "").split("(")[0].strip()
                else:
                    emp_id = emp_info or "-"
                ws.cell(row=row, column=3, value=emp_id).border = BORDER_THIN

                ws.cell(row=row, column=4, value=warning.get("field", warning.get("column", ""))).border = BORDER_THIN
                ws.cell(row=row, column=5, value=warning.get("message", warning.get("warning", ""))).border = BORDER_THIN

        # 컬럼 너비 재조정
        ws.column_dimensions['E'].width = 60


def _create_matching_sheet(ws, validation_result: Dict):
    """헤더 매칭 결과 시트"""
    matches = validation_result.get("steps", {}).get("matches", {}).get("matches", [])
    
    # 헤더
    headers = ["원본 헤더", "매칭된 필드", "신뢰도", "AI 사용", "상태"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
    
    # 데이터
    for row_idx, match in enumerate(matches, 2):
        source = match.get("source", "")
        target = match.get("target", "")
        confidence = match.get("confidence", 0)
        used_ai = match.get("used_ai", False)
        unmapped = match.get("unmapped", False)
        
        ws.cell(row=row_idx, column=1, value=source).border = BORDER_THIN
        ws.cell(row=row_idx, column=2, value=target or "(미매칭)").border = BORDER_THIN
        
        conf_cell = ws.cell(row=row_idx, column=3, value=f"{confidence * 100:.0f}%")
        conf_cell.border = BORDER_THIN
        
        ws.cell(row=row_idx, column=4, value="예" if used_ai else "아니오").border = BORDER_THIN
        
        status_cell = ws.cell(row=row_idx, column=5)
        status_cell.border = BORDER_THIN
        
        if unmapped or not target:
            status_cell.value = "❌ 미매칭"
            status_cell.fill = STYLE_ERROR
        elif confidence < 0.7:
            status_cell.value = "⚠️ 낮은 신뢰도"
            status_cell.fill = STYLE_WARNING
        else:
            status_cell.value = "✅ 매칭"
            status_cell.fill = STYLE_SUCCESS
    
    # 컬럼 너비
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def _create_layer2_sheet(ws, layer2_data: Dict):
    """인원/금액 비교 시트 (Layer 2 검증 결과)"""
    from core.ai.diagnostic_questions import get_question_by_id

    # 헤더
    headers = ["항목", "사용자 입력", "명부 계산", "차이", "차이(%)", "결과"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    # 데이터
    user_answers = layer2_data.get("user_answers", {})
    calculated = layer2_data.get("calculated_aggregates", {})
    warnings = layer2_data.get("warnings", [])

    # 경고 정보를 question_id로 매핑
    warning_map = {w.get("question_id"): w for w in warnings}

    # 비교할 질문들 (q21, q22, q23, q27, q28)
    comparison_questions = ["q21", "q22", "q23", "q27", "q28"]

    row_idx = 2
    for qid in comparison_questions:
        user_val = user_answers.get(qid)
        calc_val = calculated.get(qid)

        # 둘 다 없으면 스킵
        if user_val is None and calc_val is None:
            continue

        # 질문 텍스트 가져오기
        question = get_question_by_id(qid)
        # 질문 ID별 간단한 라벨
        label_map = {
            "q21": "임원 인원",
            "q22": "직원 인원",
            "q23": "계약직 인원",
            "q27": "퇴직금 합계",
            "q28": "중간정산 합계",
        }
        label = label_map.get(qid, question.get("question", qid)[:15] if question else qid)

        # 값 변환
        try:
            user_num = float(user_val) if user_val is not None else None
            calc_num = float(calc_val) if calc_val is not None else None
        except (ValueError, TypeError):
            user_num = None
            calc_num = None

        # 차이 계산
        if user_num is not None and calc_num is not None:
            diff = user_num - calc_num
            diff_pct = abs(diff / calc_num * 100) if calc_num != 0 else 0
            is_match = abs(diff) < 0.01 or diff_pct <= 5
        else:
            diff = None
            diff_pct = None
            is_match = None

        # 포맷팅
        unit = "명" if qid in ["q21", "q22", "q23"] else "원"
        if unit == "원" and user_num:
            user_display = f"{user_num:,.0f}원"
            calc_display = f"{calc_num:,.0f}원" if calc_num else "-"
            diff_display = f"{diff:+,.0f}원" if diff is not None else "-"
        else:
            user_display = f"{int(user_num)}명" if user_num is not None else "-"
            calc_display = f"{int(calc_num)}명" if calc_num is not None else "-"
            diff_display = f"{int(diff):+d}명" if diff is not None else "-"

        # 셀 작성
        ws.cell(row=row_idx, column=1, value=label).border = BORDER_THIN
        ws.cell(row=row_idx, column=2, value=user_display).border = BORDER_THIN
        ws.cell(row=row_idx, column=3, value=calc_display).border = BORDER_THIN
        ws.cell(row=row_idx, column=4, value=diff_display).border = BORDER_THIN

        pct_cell = ws.cell(row=row_idx, column=5, value=f"{diff_pct:.1f}%" if diff_pct is not None else "-")
        pct_cell.border = BORDER_THIN

        result_cell = ws.cell(row=row_idx, column=6)
        result_cell.border = BORDER_THIN
        result_cell.alignment = ALIGN_CENTER

        if is_match is None:
            result_cell.value = "⚪ 확인불가"
        elif is_match:
            result_cell.value = "✅ 일치"
            result_cell.fill = STYLE_SUCCESS
        else:
            result_cell.value = "❌ 불일치"
            result_cell.fill = STYLE_ERROR

        row_idx += 1

    # 컬럼 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12


def _create_anomalies_sheet(ws, validation_result: Dict):
    """이상 탐지 시트"""
    anomalies = validation_result.get("anomalies", {}).get("anomalies", [])
    
    # 헤더
    headers = ["유형", "심각도", "메시지", "필드", "값"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
    
    # 데이터
    for row_idx, anomaly in enumerate(anomalies, 2):
        ws.cell(row=row_idx, column=1, value=anomaly.get("type", "")).border = BORDER_THIN
        
        severity_cell = ws.cell(row=row_idx, column=2, value=anomaly.get("severity", ""))
        severity_cell.border = BORDER_THIN
        if anomaly.get("severity") == "high":
            severity_cell.fill = STYLE_ERROR
        elif anomaly.get("severity") == "medium":
            severity_cell.fill = STYLE_WARNING
        
        ws.cell(row=row_idx, column=3, value=anomaly.get("message", "")).border = BORDER_THIN
        ws.cell(row=row_idx, column=4, value=anomaly.get("field", "")).border = BORDER_THIN
        ws.cell(row=row_idx, column=5, value=str(anomaly.get("value", ""))).border = BORDER_THIN
    
    # 컬럼 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20


def _create_data_sheet(ws, original_data: Dict, validation_result: Dict):
    """
    원본 데이터 시트 (하이라이팅 포함)
    
    빨간색: 오류 셀 (필수값 누락, 형식 오류)
    노란색: 경고 셀 (이상치, 5% 초과 차이)
    """
    headers = original_data.get("headers", [])
    rows = original_data.get("rows", [])
    
    # 이상 탐지 정보로 하이라이팅할 셀 결정
    anomalies = validation_result.get("anomalies", {}).get("anomalies", [])
    error_cells = set()  # (row, col) 튜플
    warning_cells = set()
    
    for anomaly in anomalies:
        field = anomaly.get("field", "")
        row_idx = anomaly.get("row", None)
        severity = anomaly.get("severity", "medium")
        
        if field in headers:
            col_idx = headers.index(field)
            if row_idx is not None:
                if severity == "high":
                    error_cells.add((row_idx + 2, col_idx + 1))  # +2: 헤더 행 오프셋
                else:
                    warning_cells.add((row_idx + 2, col_idx + 1))
    
    # 헤더 작성
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) + 2)
    
    # 데이터 작성
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_THIN
            
            # 하이라이팅
            if (row_idx, col_idx) in error_cells:
                cell.fill = STYLE_ERROR
                cell.font = FONT_ERROR
            elif (row_idx, col_idx) in warning_cells:
                cell.fill = STYLE_WARNING
                cell.font = FONT_WARNING
            
            # 빈 필수값 체크
            if value is None or (isinstance(value, str) and value.strip() == ""):
                # 필수 필드인 경우 빨간색
                header = headers[col_idx - 1] if col_idx <= len(headers) else ""
                required_fields = ["사원번호", "이름", "생년월일", "입사일자", "기준급여"]
                if header in required_fields:
                    cell.fill = STYLE_ERROR
                    cell.value = "(누락)"


def export_validation_to_excel(
    validation_result: Dict[str, Any],
    original_data: Optional[Dict[str, Any]] = None,
    answers: Optional[Dict[str, Any]] = None,
    output_path: Optional[str] = None
) -> bytes:
    """
    검증 결과를 Excel 파일로 내보내기
    
    Args:
        validation_result: 검증 결과
        original_data: 원본 파싱 데이터
        answers: 진단 답변
        output_path: 저장 경로 (없으면 bytes 반환)
    
    Returns:
        Excel 파일 bytes
    """
    excel_bytes = generate_excel_report(validation_result, original_data, answers)
    
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(excel_bytes)
    
    return excel_bytes


def generate_final_data_excel(
    original_data: Dict[str, Any],
    validation_result: Dict[str, Any]
) -> bytes:
    """
    최종 수정본 Excel 생성 (매핑 완료된 깔끔한 데이터)
    
    원본 헤더를 표준 필드명으로 변환하여 정리된 데이터 제공
    
    Args:
        original_data: 원본 파싱 데이터 (headers, rows)
        validation_result: 검증 결과 (매핑 정보 포함)
    
    Returns:
        Excel 파일 bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "최종 데이터"
    
    # 매핑 정보 추출
    matches = validation_result.get("steps", {}).get("matches", {}).get("matches", [])
    
    # 원본 헤더 → 표준 필드명 매핑
    header_mapping = {}
    for match in matches:
        source = match.get("source", "")
        target = match.get("target")
        if target:  # 매핑된 것만
            header_mapping[source] = target
    
    original_headers = original_data.get("headers", [])
    rows = original_data.get("rows", [])
    
    # 새 헤더 (표준 필드명으로 변환)
    new_headers = []
    included_cols = []  # 포함할 컬럼 인덱스
    
    for idx, header in enumerate(original_headers):
        if header in header_mapping:
            new_headers.append(header_mapping[header])
            included_cols.append(idx)
        else:
            # 미매핑 컬럼도 포함 (원본 이름 유지)
            new_headers.append(f"[미매핑] {header}")
            included_cols.append(idx)
    
    # 헤더 작성
    for col_idx, header in enumerate(new_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        
        if header.startswith("[미매핑]"):
            cell.fill = STYLE_WARNING
            cell.font = FONT_WARNING
        else:
            cell.fill = STYLE_HEADER
            cell.font = FONT_HEADER
        
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(header)) + 2)
    
    # 데이터 작성
    for row_idx, row_data in enumerate(rows, 2):
        for new_col_idx, orig_col_idx in enumerate(included_cols, 1):
            value = row_data[orig_col_idx] if orig_col_idx < len(row_data) else ""
            cell = ws.cell(row=row_idx, column=new_col_idx, value=value)
            cell.border = BORDER_THIN
            
            # 빈 필수값 하이라이트
            if value is None or (isinstance(value, str) and value.strip() == ""):
                header = new_headers[new_col_idx - 1]
                required_fields = ["사원번호", "이름", "생년월일", "입사일자", "기준급여"]
                if header in required_fields:
                    cell.fill = STYLE_ERROR
                    cell.value = "(누락)"
    
    # 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
