"""
Excel Exporter for Validation Results with Color Coding and Comments
"""
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment


def export_validation_results(parsed_data: Dict[str, Any], output_path: str) -> str:
    """
    Export validation results to Excel with color coding and comments
    
    Colors:
    - Red: Errors (status='error', 'mismatch' with high impact)
    - Yellow: Warnings (status='warning', 'mismatch' with low impact)
    - Green: Match (status='match')
    
    Args:
        parsed_data: Dictionary from ceragem_parser.parse_all()
        output_path: Path to save Excel file
        
    Returns:
        Path to exported file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Color definitions
    RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    GRAY_FILL = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    BOLD_FONT = Font(bold=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet('검증 요약')
    ws_summary.append(['항목', '값', '상태', '비고'])
    
    # Apply header formatting
    for col in range(1, 5):
        cell = ws_summary.cell(1, col)
        cell.font = BOLD_FONT
        cell.fill = GRAY_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Cross-checks summary
    cross_checks = parsed_data.get('cross_checks', [])
    ws_summary.append(['', '', '', ''])
    ws_summary.append(['교차 검증 결과', '', '', ''])
    ws_summary.cell(ws_summary.max_row, 1).font = BOLD_FONT
    
    for check in cross_checks:
        check_type = check.get('type', '')
        status = check.get('status', '')
        label = check.get('label', check_type)
        
        # Determine fill color
        if status == 'match':
            fill = GREEN_FILL
            status_text = '✓ 일치'
        elif status == 'mismatch':
            diff = check.get('diff', 0)
            if abs(diff) > 1000000:  # High impact
                fill = RED_FILL
                status_text = '✗ 불일치 (확인필요)'
            else:
                fill = YELLOW_FILL
                status_text = '△ 불일치 (검토필요)'
        elif status == 'error':
            fill = RED_FILL
            status_text = '✗ 오류'
        else:
            fill = None
            status_text = status
        
        # Build remark
        remark = ''
        if 'formula_' in check_type:
            lhs = check.get('lhs', 0)
            rhs = check.get('rhs', 0)
            diff = check.get('diff', 0)
            remark = f"LHS={lhs:,.0f}, RHS={rhs:,.0f}, Diff={diff:,.0f}"
        elif 'compare_' in check_type:
            core = check.get('core_total', 0)
            lists = check.get('lists_total', 0)
            diff = check.get('diff', 0)
            remark = f"Core={core:,.0f}, Lists={lists:,.0f}, Diff={diff:,.0f}"
        
        row_num = ws_summary.max_row + 1
        ws_summary.append([label, '', status_text, remark])
        
        # Apply formatting
        for col in range(1, 5):
            cell = ws_summary.cell(row_num, col)
            cell.border = THIN_BORDER
            if fill and col in [2, 3]:
                cell.fill = fill
            if col == 3:
                cell.alignment = CENTER_ALIGN
        
        # Add comment for mismatches
        if status == 'mismatch' and remark:
            comment = Comment(f"불일치 발견:\n{remark}\n\n고객 확인 필요", "System")
            ws_summary.cell(row_num, 3).comment = comment
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 50
    
    # Sheet 2: Formula Validations
    ws_formulas = wb.create_sheet('공식 검증')
    ws_formulas.append(['공식', '계산값 (LHS)', '저장값 (RHS)', '차이', '상태'])
    
    for col in range(1, 6):
        cell = ws_formulas.cell(1, col)
        cell.font = BOLD_FONT
        cell.fill = GRAY_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    for check in cross_checks:
        if 'formula_' not in check.get('type', ''):
            continue
        
        label = check.get('label', check.get('type', ''))
        lhs = check.get('lhs', 0)
        rhs = check.get('rhs', 0)
        diff = check.get('diff', 0)
        status = check.get('status', '')
        
        row_num = ws_formulas.max_row + 1
        ws_formulas.append([label, lhs, rhs, diff, status])
        
        # Format numbers
        for col in [2, 3, 4]:
            ws_formulas.cell(row_num, col).number_format = '#,##0.00'
        
        # Color based on status
        fill = GREEN_FILL if status == 'match' else (RED_FILL if abs(diff) > 1e-3 else YELLOW_FILL)
        for col in range(1, 6):
            cell = ws_formulas.cell(row_num, col)
            cell.border = THIN_BORDER
            if col in [4, 5]:
                cell.fill = fill
        
        # Add comment if mismatch
        if status != 'match':
            comment_text = f"공식 불일치:\n계산값: {lhs:,.2f}\n저장값: {rhs:,.2f}\n차이: {diff:,.2f}\n\n원인 확인 필요"
            ws_formulas.cell(row_num, 5).comment = Comment(comment_text, "System")
    
    ws_formulas.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D']:
        ws_formulas.column_dimensions[col].width = 18
    ws_formulas.column_dimensions['E'].width = 15
    
    # Sheet 3: Labeled Cells (I40-I51)
    ws_cells = wb.create_sheet('지정 셀 값')
    ws_cells.append(['셀 주소', '라벨', '값', '비고'])
    
    for col in range(1, 5):
        cell = ws_cells.cell(1, col)
        cell.font = BOLD_FONT
        cell.fill = GRAY_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    from internal.parsers.ceragem_parser import CELL_I40_I51_LABELS
    sums = parsed_data.get('core_info', {}).get('sums_I40_I51', [])
    
    for i, (cell_addr, label) in enumerate(CELL_I40_I51_LABELS.items()):
        value = sums[i] if i < len(sums) else ''
        row_num = ws_cells.max_row + 1
        ws_cells.append([cell_addr, label, value, ''])
        
        # Format value
        if isinstance(value, (int, float)) and value != '':
            ws_cells.cell(row_num, 3).number_format = '#,##0'
        
        for col in range(1, 5):
            ws_cells.cell(row_num, col).border = THIN_BORDER
    
    ws_cells.column_dimensions['A'].width = 12
    ws_cells.column_dimensions['B'].width = 40
    ws_cells.column_dimensions['C'].width = 20
    ws_cells.column_dimensions['D'].width = 30
    
    # Save workbook
    wb.save(output_path)
    return output_path


def create_validation_report(parsed_data: Dict, validation_results: Dict, output_path: str) -> str:
    """
    Create comprehensive validation report combining parse results and Layer 1 validation
    
    Args:
        parsed_data: From ceragem_parser.parse_all()
        validation_results: From validation_layer1.validate_layer1()
        output_path: Path to save Excel
        
    Returns:
        Path to exported file
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Color definitions
    RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    BOLD_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Validation Errors
    ws_errors = wb.create_sheet('검증 오류')
    ws_errors.append(['행번호', '컬럼', '현재값', '오류 메시지', '심각도'])
    
    for col in range(1, 6):
        cell = ws_errors.cell(1, col)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    
    errors = validation_results.get('errors', [])
    for err in errors[:500]:  # Limit to first 500
        row = err.get('row', '')
        col = err.get('column', '')
        val = err.get('value', '')
        msg = err.get('message', '')
        severity = '오류'
        
        row_num = ws_errors.max_row + 1
        ws_errors.append([row, col, val, msg, severity])
        
        for c in range(1, 6):
            cell = ws_errors.cell(row_num, c)
            cell.border = THIN_BORDER
            if c == 5:
                cell.fill = RED_FILL
    
    ws_errors.column_dimensions['A'].width = 10
    ws_errors.column_dimensions['B'].width = 20
    ws_errors.column_dimensions['C'].width = 20
    ws_errors.column_dimensions['D'].width = 50
    ws_errors.column_dimensions['E'].width = 12
    
    # Sheet 2: Warnings
    ws_warnings = wb.create_sheet('검증 경고')
    ws_warnings.append(['행번호', '컬럼', '현재값', '경고 메시지', '심각도'])
    
    for col in range(1, 6):
        cell = ws_warnings.cell(1, col)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    
    warnings = validation_results.get('warnings', [])
    for warn in warnings[:500]:
        row = warn.get('row', '')
        col = warn.get('column', '')
        val = warn.get('value', '')
        msg = warn.get('message', '')
        severity = '경고'
        
        row_num = ws_warnings.max_row + 1
        ws_warnings.append([row, col, val, msg, severity])
        
        for c in range(1, 6):
            cell = ws_warnings.cell(row_num, c)
            cell.border = THIN_BORDER
            if c == 5:
                cell.fill = YELLOW_FILL
    
    ws_warnings.column_dimensions['A'].width = 10
    ws_warnings.column_dimensions['B'].width = 20
    ws_warnings.column_dimensions['C'].width = 20
    ws_warnings.column_dimensions['D'].width = 50
    ws_warnings.column_dimensions['E'].width = 12
    
    wb.save(output_path)
    return output_path
