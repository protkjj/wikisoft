"""
(1-2) 기초자료_퇴직급여 시트 생성기

챗봇 답변 기반으로 (1-2) 시트 생성
Layer 2 검증 결과에서 경고가 있으면 해당 셀에 빨간 배경 + 코멘트 추가
"""

from typing import Dict, Any, Optional, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from datetime import datetime
import io


def create_sheet_1_2_from_chatbot(
    chatbot_answers: Dict[str, Any],
    validation_warnings: List[Dict[str, Any]] = None,
    company_info: Optional[Dict[str, str]] = None,
    작성기준일: str = None
) -> Workbook:
    """
    챗봇 답변으로 (1-2) 기초자료_퇴직급여 시트 생성
    
    Args:
        chatbot_answers: 챗봇 답변 딕셔너리 {"q15": 3.5, "q21": 17, ...}
        validation_warnings: Layer 2 검증 경고 리스트
        company_info: 회사 정보
        작성기준일: YYYYMMDD
        
    Returns:
        Workbook 객체 (검증 경고는 빨간 배경 + 코멘트 표시)
    """
    from internal.ai.diagnostic_questions import ALL_QUESTIONS
    
    wb = Workbook()
    ws = wb.active
    ws.title = "(1-2) 기초자료_퇴직급여"
    
    # 스타일 정의
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    label_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    warning_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')  # 노란색
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 빨간색
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 경고 맵핑: question_id → cell_address
    # (나중에 셀 주소를 동적으로 계산하여 사용)
    warnings_by_qid = {}
    if validation_warnings:
        for warning in validation_warnings:
            qid = warning.get('question_id')
            warnings_by_qid[qid] = warning
    
    # 제목
    ws['A1'] = '(1-2) 퇴직급여채무 평가를 위한 기초자료'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:K1')
    
    # 회사 정보
    if company_info is None:
        company_info = {}
    
    info_start_row = 3
    ws[f'C{info_start_row}'] = '회사정보'
    ws[f'C{info_start_row}'].font = bold_font
    ws[f'C{info_start_row}'].fill = header_fill
    
    # 인원수 집계 섹션
    count_start_row = 12
    ws[f'H{count_start_row}'] = '평가 대상 인원수'
    ws[f'H{count_start_row}'].font = bold_font
    ws[f'H{count_start_row}'].fill = header_fill
    
    # 챗봇 답변 매핑 (question_id → cell)
    cell_mapping = {
        # 재무적 가정
        'q15': 'I62',  # 할인율
        'q16': 'I68',  # 승급률
        'q17': 'I74',  # 임금상승률
        # 퇴직금 설정
        'q18': 'F103',  # 퇴직금 한도
        'q19': 'D121',  # 평균근속연수 추정
        'q20': 'F126',  # 임원 제외 여부
        # 인원 집계
        'q21': 'I14',  # 재직 임원 (I26에 해당, 상대 행번호 조정)
        'q22': 'I15',  # 재직 직원
        'q23': 'I16',  # 재직 계약직
        'q24': 'I21',  # 퇴직자 합계
        'q25': 'I22',  # 중간정산자
        # 금액 집계
        'q26': 'I29',  # 퇴직금 총액 (I41 상대 행번호)
        'q27': 'I30',  # 중간정산 금액
        'q28': 'I31',  # DC전환금
    }
    
    # 값 채우기 및 검증 경고 표시
    for qid, cell_addr in cell_mapping.items():
        value = chatbot_answers.get(qid)
        if value is not None:
            ws[cell_addr] = value
            ws[cell_addr].number_format = '#,##0' if isinstance(value, (int, float)) else '@'
            
            # 검증 경고가 있으면 스타일 적용
            if qid in warnings_by_qid:
                warning = warnings_by_qid[qid]
                severity = warning.get('severity', 'low')
                
                if severity == 'high':
                    ws[cell_addr].fill = error_fill
                    comment_text = f"⭕ {warning.get('message', '')}\n계산값: {warning.get('calculated')}"
                elif severity == 'low':
                    ws[cell_addr].fill = warning_fill
                    comment_text = f"⚠️ {warning.get('message', '')}"
                else:
                    comment_text = warning.get('message', '')
                
                # 코멘트 추가
                ws[cell_addr].comment = Comment(comment_text, "WIKISOFT2 검증")
    
    # 라벨 추가 (간소화 버전)
    ws['H14'] = '임원'
    ws['H15'] = '직원'
    ws['H16'] = '계약직'
    ws['H17'] = '합계'
    ws['H21'] = '퇴직자 합계'
    ws['H22'] = '중간정산자'
    
    ws['D29'] = '퇴직금 (a)'
    ws['D30'] = '중간정산금액 (b)'
    ws['D31'] = 'DC 전환금 (c)'
    
    return wb
    """
    (1-2) 기초자료_퇴직급여 시트 생성
    
    Args:
        aggregated: aggregate_calculator의 결과
        company_info: 회사 정보 {'회사명', '전화번호', '휴대폰', '이메일', '결산일', '주소'}
        작성기준일: YYYYMMDD 형식
        
    Returns:
        Workbook 객체
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "(1-2) 기초자료_퇴직급여"
    
    # 스타일 정의
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    label_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    bold_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 제목
    ws['A1'] = '(1-2) 퇴직급여채무 평가를 위한 기초자료'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:K1')
    
    # 회사 정보 섹션 (D12-D17 매핑)
    if company_info is None:
        company_info = {}
    
    info_start_row = 3
    ws[f'C{info_start_row}'] = '회사정보'
    ws[f'C{info_start_row}'].font = bold_font
    ws[f'C{info_start_row}'].fill = header_fill
    
    info_items = [
        ('D', info_start_row + 1, '회사명', company_info.get('회사명', '')),
        ('D', info_start_row + 2, '전화번호', company_info.get('전화번호', '')),
        ('D', info_start_row + 3, '휴대폰', company_info.get('휴대폰', '')),
        ('D', info_start_row + 4, '이메일', company_info.get('이메일', '')),
        ('D', info_start_row + 5, '결산일', company_info.get('결산일', '')),
        ('D', info_start_row + 6, '주소', company_info.get('주소', '')),
    ]
    
    for col, row, label, value in info_items:
        ws[f'C{row}'] = label
        ws[f'{col}{row}'] = value
        ws[f'C{row}'].fill = label_fill
    
    # 인원수 집계 섹션 (I25-I39)
    count_start_row = 12
    ws[f'H{count_start_row}'] = '평가 대상 인원수'
    ws[f'H{count_start_row}'].font = bold_font
    ws[f'H{count_start_row}'].fill = header_fill
    ws.merge_cells(f'H{count_start_row}:I{count_start_row}')
    
    # I25: 작성기준일
    ws[f'H{count_start_row + 1}'] = '작성기준일'
    ws[f'I{count_start_row + 1}'] = 작성기준일 if 작성기준일 else datetime.now().strftime('%Y%m%d')
    
    # I26-I29: 재직자
    counts = aggregated['counts_I26_I39']
    ws[f'D{count_start_row + 2}'] = '재직자수(명)'
    ws[f'H{count_start_row + 2}'] = '임원'
    ws[f'I{count_start_row + 2}'] = counts[0] if counts[0] is not None else 0  # I26
    ws[f'H{count_start_row + 3}'] = '직원'
    ws[f'I{count_start_row + 3}'] = counts[1] if counts[1] is not None else 0  # I27
    ws[f'H{count_start_row + 4}'] = '계약직'
    ws[f'I{count_start_row + 4}'] = counts[2] if counts[2] is not None else 0  # I28
    ws[f'H{count_start_row + 5}'] = '합계'
    ws[f'I{count_start_row + 5}'] = counts[3] if counts[3] is not None else 0  # I29
    
    # I30-I33: 퇴직자
    ws[f'D{count_start_row + 6}'] = '퇴직자수(명)'
    ws[f'H{count_start_row + 6}'] = '임원'
    ws[f'I{count_start_row + 6}'] = counts[4] if counts[4] is not None else 0  # I30
    ws[f'H{count_start_row + 7}'] = '직원'
    ws[f'I{count_start_row + 7}'] = counts[5] if counts[5] is not None else 0  # I31
    ws[f'H{count_start_row + 8}'] = '계약직'
    ws[f'I{count_start_row + 8}'] = counts[6] if counts[6] is not None else 0  # I32
    ws[f'H{count_start_row + 9}'] = '합계'
    ws[f'I{count_start_row + 9}'] = counts[7] if counts[7] is not None else 0  # I33
    
    # I34-I35: 중간정산, DC전환
    ws[f'D{count_start_row + 10}'] = '중간정산자수'
    ws[f'I{count_start_row + 10}'] = counts[8] if counts[8] is not None else 0  # I34
    ws[f'D{count_start_row + 11}'] = 'DC전환자수'
    ws[f'I{count_start_row + 11}'] = counts[9] if counts[9] is not None else 0  # I35
    
    # 금액 집계 섹션 (I40-I51)
    amount_start_row = 30
    ws[f'B{amount_start_row}'] = '퇴직부채 감소(증가)액(원)'
    ws[f'B{amount_start_row}'].font = bold_font
    ws[f'B{amount_start_row}'].fill = header_fill
    
    sums = aggregated['sums_I40_I51']
    
    # I41-I44: 급여지급액
    ws[f'C{amount_start_row + 1}'] = '급여지급액(A)'
    ws[f'D{amount_start_row + 1}'] = '퇴직금 (a)'
    ws[f'I{amount_start_row + 1}'] = sums[1] if sums[1] is not None else 0  # I41
    ws[f'D{amount_start_row + 2}'] = '중간정산금액 (b)'
    ws[f'I{amount_start_row + 2}'] = sums[2] if sums[2] is not None else 0  # I42
    ws[f'D{amount_start_row + 3}'] = 'DC 전환금 (c)'
    ws[f'I{amount_start_row + 3}'] = sums[3] if sums[3] is not None else 0  # I43
    ws[f'D{amount_start_row + 4}'] = '소계 (a+b+c)'
    ws[f'I{amount_start_row + 4}'] = f'=I{amount_start_row + 1}+I{amount_start_row + 2}+I{amount_start_row + 3}'
    
    # I45-I47: 관계사 전입/전출
    ws[f'C{amount_start_row + 5}'] = '관계사전입(전출)액(B)'
    ws[f'D{amount_start_row + 5}'] = '관계사 전입액 (d)'
    ws[f'I{amount_start_row + 5}'] = sums[5] if sums[5] is not None else 0  # I45
    ws[f'D{amount_start_row + 6}'] = '관계사 전출액 (e)'
    ws[f'I{amount_start_row + 6}'] = sums[6] if sums[6] is not None else 0  # I46
    ws[f'D{amount_start_row + 7}'] = '소계 (d-e)'
    ws[f'I{amount_start_row + 7}'] = f'=I{amount_start_row + 5}-I{amount_start_row + 6}'
    
    # I48-I50: 사업결합/처분
    ws[f'C{amount_start_row + 8}'] = '사업결합(처분)액(C)'
    ws[f'D{amount_start_row + 8}'] = '사업결합액 (f)'
    ws[f'I{amount_start_row + 8}'] = sums[8] if sums[8] is not None else 0  # I48
    ws[f'D{amount_start_row + 9}'] = '사업처분액 (g)'
    ws[f'I{amount_start_row + 9}'] = sums[9] if sums[9] is not None else 0  # I49
    ws[f'D{amount_start_row + 10}'] = '소계 (f-g)'
    ws[f'I{amount_start_row + 10}'] = f'=I{amount_start_row + 8}-I{amount_start_row + 9}'
    
    # I51: 최종 합계
    ws[f'B{amount_start_row + 11}'] = '퇴직부채 감소(증가)액 합계 (A-B-C)'
    ws[f'I{amount_start_row + 11}'] = f'=I{amount_start_row + 4}-I{amount_start_row + 7}-I{amount_start_row + 10}'
    
    # 숫자 포맷 적용
    for row in range(amount_start_row + 1, amount_start_row + 12):
        ws[f'I{row}'].number_format = '#,##0'
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    
    return wb


def save_sheet_1_2_to_bytes(
    aggregated: Dict[str, Any],
    company_info: Optional[Dict[str, str]] = None,
    작성기준일: str = None
) -> bytes:
    """
    (1-2) 시트를 생성해서 bytes로 반환 (다운로드용)
    
    Returns:
        Excel 파일의 bytes
    """
    wb = create_sheet_1_2(aggregated, company_info, 작성기준일)
    
    # BytesIO로 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def save_sheet_1_2_to_file(
    aggregated: Dict[str, Any],
    output_path: str,
    company_info: Optional[Dict[str, str]] = None,
    작성기준일: str = None
):
    """
    (1-2) 시트를 파일로 저장
    
    Args:
        output_path: 저장할 파일 경로
    """
    wb = create_sheet_1_2(aggregated, company_info, 작성기준일)
    wb.save(output_path)
    return output_path


def save_sheet_1_2_from_chatbot_to_bytes(
    chatbot_answers: Dict[str, Any],
    validation_warnings: List[Dict[str, Any]] = None,
    company_info: Optional[Dict[str, str]] = None,
    작성기준일: str = None
) -> bytes:
    """
    챗봇 답변 기반 (1-2) 시트 생성하여 bytes 반환 (다운로드용)
    
    Returns:
        Excel 파일 bytes
    """
    wb = create_sheet_1_2_from_chatbot(chatbot_answers, validation_warnings, company_info, 작성기준일)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
