"""
Layer 2 검증 시스템 통합 테스트

시나리오:
1. 명부 파일 업로드
2. 챗봇 28개 질문에 답변 (일부러 틀린 답변 포함)
3. Layer 2 검증 실행 (명부 계산값 vs 챗봇 답변)
4. 검증 경고가 포함된 Excel 파일 생성
"""
import sys
sys.path.insert(0, "/Users/kj/Desktop/wiki/WIKISOFT2")

from internal.generators.aggregate_calculator import aggregate_from_excel
from internal.processors.validation_layer2 import validate_chatbot_answers
from internal.generators.sheet_generator import create_sheet_1_2_from_chatbot
from openpyxl.styles.fills import PatternFill
import json


def test_full_workflow():
    print("=" * 70)
    print("Layer 2 검증 시스템 통합 테스트")
    print("=" * 70)
    
    # ========== 1. 명부 파일 읽기 ==========
    print("\n[1단계] 명부 파일 파싱...")
    with open('20251223_세라젬_202512_확정급여채무평가_작성요청자료_기타장기 제외_메일발송.xls', 'rb') as f:
        roster_content = f.read()
    
    # ========== 2. 자동 집계 계산 ==========
    print("[2단계] 명부에서 자동 집계 계산...")
    calculated = aggregate_from_excel(roster_content)
    
    print(f"  ✓ 계산 완료")
    print(f"    - 임원: {calculated['counts_I26_I39'][0]:.0f}명")
    print(f"    - 직원: {calculated['counts_I26_I39'][1]:.0f}명")
    print(f"    - 계약직: {calculated['counts_I26_I39'][2]:.0f}명")
    print(f"    - 퇴직자: {calculated['counts_I26_I39'][7]:.0f}명")
    
    # ========== 3. 챗봇 답변 시뮬레이션 (일부러 틀리게) ==========
    print("\n[3단계] 챗봇 답변 시뮬레이션...")
    chatbot_answers = {
        # 데이터 품질 (q1~q14) - 생략
        # 재무적 가정
        "q15": 3.5,   # 할인율
        "q16": 2.0,   # 승급률
        "q17": 3.0,   # 임금상승률
        # 퇴직금 설정
        "q18": 60,    # 퇴직금 한도
        "q19": "회사채AA+",  # 평균근속연수 추정
        "q20": "제외",       # 임원 제외 여부
        # 인원 집계 (일부러 틀리게)
        "q21": 20,    # 임원 - 실제 17, 틀림! 🔴
        "q22": 664,   # 직원 - 정확 ✓
        "q23": 69,    # 계약직 - 정확 ✓
        "q24": 480,   # 퇴직자 - 실제 477, 소폭 차이 🟡
        "q25": 26,    # 중간정산자 - 정확 ✓
        # 금액 집계 (계산 불가 항목이므로 사용자 입력 신뢰)
        "q26": 6800000000,  # 퇴직금 68억
        "q27": 691876810,   # 중간정산
        "q28": 0            # DC전환금
    }
    
    print(f"  ✓ 총 {len(chatbot_answers)}개 답변 준비")
    print(f"    - 재무적 가정: 3개")
    print(f"    - 퇴직금 설정: 3개")
    print(f"    - 인원 집계: 5개")
    print(f"    - 금액 집계: 3개")
    
    # ========== 4. Layer 2 검증 실행 ==========
    print("\n[4단계] Layer 2 검증 실행...")
    validation = validate_chatbot_answers(
        chatbot_answers,
        calculated,
        tolerance_percent=5.0
    )
    
    print(f"  상태: {validation['status'].upper()}")
    print(f"  검사 항목: {validation['total_checks']}개")
    print(f"  통과: {validation['passed']}개")
    print(f"  경고: {len(validation['warnings'])}개")
    
    # 경고 상세
    if validation['warnings']:
        print("\n  📋 경고 상세:")
        for w in validation['warnings']:
            severity_icon = {"high": "🔴", "low": "🟡", "info": "ℹ️"}.get(w['severity'], '⚪')
            print(f"    {severity_icon} [{w['question_id']}] {w['severity'].upper()}")
            print(f"       {w['message'][:80]}")
    
    # ========== 5. Excel 파일 생성 ==========
    print("\n[5단계] 검증 결과 포함 Excel 생성...")
    
    wb = create_sheet_1_2_from_chatbot(
        chatbot_answers,
        validation_warnings=validation['warnings'],
        company_info={
            '회사명': '테스트 주식회사',
            '전화번호': '02-1234-5678'
        },
        작성기준일='20251231'
    )
    
    output_path = 'test_layer2_validation.xlsx'
    wb.save(output_path)
    
    print(f"  ✓ 파일 생성: {output_path}")
    
    # ========== 6. 생성된 파일 검증 ==========
    print("\n[6단계] 생성된 Excel 검증...")
    ws = wb.active
    
    # q21 (임원 20명, 실제 17명) → I14 셀 검증
    cell_i14 = ws['I14']
    print(f"  I14 (임원): {cell_i14.value}")
    if cell_i14.fill and cell_i14.fill.start_color.rgb:
        fill_color = cell_i14.fill.start_color.rgb
        if fill_color in ['FFFFC7CE', 'FFC7CE']:  # 빨간색
            print(f"    ✓ 빨간 배경 표시됨 (색상: {fill_color})")
    if cell_i14.comment:
        print(f"    ✓ 코멘트: {cell_i14.comment.text[:50]}...")
    
    # ========== 7. JSON 응답 시뮬레이션 ==========
    print("\n[7단계] API 응답 형식 (JSON)...")
    api_response = {
        "validation": {
            "status": validation['status'],
            "total_checks": validation['total_checks'],
            "passed": validation['passed'],
            "warnings": [
                {
                    "question_id": w['question_id'],
                    "severity": w['severity'],
                    "message": w['message']
                }
                for w in validation['warnings']
            ]
        },
        "excel_file": output_path,
        "message": "검증 완료. Excel 파일을 다운로드하세요."
    }
    
    print(json.dumps(api_response, ensure_ascii=False, indent=2))
    
    # ========== 8. 최종 요약 ==========
    print("\n" + "=" * 70)
    print("✅ 통합 테스트 완료")
    print("=" * 70)
    print(f"검증 상태: {validation['status'].upper()}")
    if validation['status'] == 'failed':
        print("⚠️  심각한 불일치가 발견되었습니다. Excel 파일을 확인하세요.")
    elif validation['status'] == 'warnings':
        print("⚠️  경미한 차이가 있습니다. 확인이 필요합니다.")
    else:
        print("✅ 모든 항목이 일치합니다.")
    
    print(f"\n생성된 파일: {output_path}")
    print("빨간 배경 셀 = 심각한 불일치, 노란 배경 셀 = 경미한 차이")
    print("셀에 마우스를 올리면 상세 메시지를 확인할 수 있습니다.")


if __name__ == "__main__":
    test_full_workflow()
