"""
Windmill 워크플로우 통합 엔드포인트

- /trigger: 워크플로우 트리거 (파일 검증 → 분기 결정)
- /callback: 실행 결과 수신 및 로깅
- /latest: 최근 실행 기록 조회
"""
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
from fastapi import APIRouter, File, Form, HTTPException, Header, UploadFile
from pydantic import BaseModel

from internal.agent.tool_registry import get_registry
from internal.agent.confidence import detect_anomalies, estimate_confidence
from internal.utils.security import validate_upload_file, secure_logger

router = APIRouter(prefix="/api/windmill", tags=["windmill"])

# 환경변수에서 시크릿 로드
WINDMILL_CALLBACK_SECRET = os.getenv("WINDMILL_CALLBACK_SECRET", "")

# 자동 승인 기준
AUTO_APPROVE_CONFIDENCE_THRESHOLD = float(os.getenv("AUTO_APPROVE_THRESHOLD", "0.90"))


# ============================================================
# 워크플로우 트리거 응답 모델
# ============================================================
class WorkflowResult(BaseModel):
    """워크플로우 트리거 결과"""
    success: bool
    action: str  # "auto_approve" | "needs_review" | "rejected"
    auto_approve: bool
    confidence: float
    message: str

    # 상세 결과 (Windmill에서 다음 단계로 전달)
    parsed: Optional[Dict[str, Any]] = None
    matches: Optional[Dict[str, Any]] = None
    validation: Optional[Dict[str, Any]] = None
    anomalies: Optional[Dict[str, Any]] = None
    duplicates: Optional[Dict[str, Any]] = None

    # 에러 정보
    errors: Optional[List[Dict[str, Any]]] = None
    warnings: Optional[List[Dict[str, Any]]] = None


# ============================================================
# 워크플로우 트리거 엔드포인트
# ============================================================
@router.post("/trigger")
async def trigger_workflow(
    file: UploadFile = File(...),
    sheet_type: str = Form("재직자"),
    auto_approve_threshold: Optional[float] = Form(None),
    chatbot_answers: Optional[str] = Form(None)
) -> WorkflowResult:
    """
    Windmill 워크플로우 트리거 엔드포인트.

    파일 검증 후 auto_approve / needs_review / rejected 분기 결정.

    Input:
        - file: Excel 명부 파일
        - sheet_type: 시트 유형 (재직자/퇴직자)
        - auto_approve_threshold: 자동 승인 신뢰도 기준 (기본 0.90)
        - chatbot_answers: 진단 질문 답변 (JSON 문자열)

    Returns:
        - action: "auto_approve" | "needs_review" | "rejected"
        - auto_approve: 자동 승인 가능 여부
        - confidence: 신뢰도 점수
        - message: 결과 메시지
        - 상세 결과 (parsed, matches, validation, anomalies, duplicates)

    Windmill 플로우에서 사용 예시:
    ```
    result = POST /api/windmill/trigger (file, sheet_type)

    if result.action == "auto_approve":
        → 자동 승인 → IFRS 계산 진행
    elif result.action == "needs_review":
        → 담당자에게 알림 → 수동 검토 대기
    else:  # rejected
        → 파일 반려 → 사용자에게 재업로드 요청
    ```
    """
    threshold = auto_approve_threshold or AUTO_APPROVE_CONFIDENCE_THRESHOLD

    try:
        # 파일 검증
        file_bytes, filename = await validate_upload_file(file)
        registry = get_registry()

        # 1. 파싱
        parsed = registry.call_tool("parse_roster", file_bytes=file_bytes)
        if not parsed.get("headers") or not parsed.get("rows"):
            return WorkflowResult(
                success=False,
                action="rejected",
                auto_approve=False,
                confidence=0.0,
                message="파일 파싱 실패: 헤더 또는 데이터가 없습니다.",
                errors=[{"type": "parse_error", "message": "빈 파일이거나 형식이 잘못되었습니다."}]
            )

        # 2. 헤더 매칭
        matches = registry.call_tool("match_headers", parsed=parsed, sheet_type=sheet_type)

        # 3. 진단 답변 파싱
        diagnostic_answers = {}
        if chatbot_answers:
            try:
                diagnostic_answers = json.loads(chatbot_answers)
            except json.JSONDecodeError:
                pass

        # 4. 검증
        validation = registry.call_tool(
            "validate",
            parsed=parsed,
            matches=matches,
            diagnostic_answers=diagnostic_answers
        )

        # 5. 신뢰도 및 이상치 분석
        confidence_result = estimate_confidence(parsed, matches, validation)
        anomalies_result = detect_anomalies(parsed, matches, validation)

        # 6. 중복 탐지
        df = pd.DataFrame(parsed.get("rows", []), columns=parsed.get("headers", []))
        duplicates = registry.call_tool(
            "detect_duplicates",
            df=df,
            headers=parsed.get("headers", []),
            matches=matches.get("matches", [])
        )

        # 7. 분기 결정
        confidence_score = confidence_result.get("score", 0)
        errors = [a for a in anomalies_result.get("anomalies", []) if a.get("severity") == "error"]
        warnings = [a for a in anomalies_result.get("anomalies", []) if a.get("severity") == "warning"]

        # 검증 결과에서도 에러/경고 추출
        validation_errors = validation.get("errors", [])
        validation_warnings = validation.get("warnings", [])

        all_errors = errors + validation_errors
        all_warnings = warnings + validation_warnings

        # 분기 로직
        action, auto_approve, message = _decide_action(
            confidence_score=confidence_score,
            threshold=threshold,
            errors=all_errors,
            warnings=all_warnings,
            duplicates=duplicates
        )

        return WorkflowResult(
            success=True,
            action=action,
            auto_approve=auto_approve,
            confidence=confidence_score,
            message=message,
            parsed={
                "headers": parsed.get("headers", []),
                "row_count": len(parsed.get("rows", [])),
                "rows": parsed.get("rows", [])  # Windmill에서 다음 단계로 전달
            },
            matches=matches,
            validation=validation,
            anomalies=anomalies_result,
            duplicates=duplicates,
            errors=all_errors if all_errors else None,
            warnings=all_warnings if all_warnings else None
        )

    except Exception as e:
        return WorkflowResult(
            success=False,
            action="rejected",
            auto_approve=False,
            confidence=0.0,
            message=f"처리 중 오류 발생: {str(e)}",
            errors=[{"type": "system_error", "message": str(e)}]
        )


def _decide_action(
    confidence_score: float,
    threshold: float,
    errors: List[Dict],
    warnings: List[Dict],
    duplicates: Dict
) -> tuple:
    """
    auto_approve / needs_review / rejected 분기 결정

    Returns:
        (action, auto_approve, message)
    """
    has_errors = len(errors) > 0
    has_critical_duplicates = len(duplicates.get("exact_duplicates", [])) > 0
    has_warnings = len(warnings) > 0

    # Case 1: 에러가 있으면 rejected
    if has_errors:
        error_summary = ", ".join([e.get("message", "알 수 없는 오류")[:30] for e in errors[:3]])
        return (
            "rejected",
            False,
            f"검증 실패: {len(errors)}개 에러 발견. ({error_summary})"
        )

    # Case 2: 완전 중복이 있으면 needs_review
    if has_critical_duplicates:
        return (
            "needs_review",
            False,
            f"중복 데이터 발견: {len(duplicates.get('exact_duplicates', []))}건의 완전 중복. 수동 검토 필요."
        )

    # Case 3: 신뢰도가 기준 이상이고 경고만 있으면 auto_approve
    if confidence_score >= threshold:
        if has_warnings:
            return (
                "auto_approve",
                True,
                f"자동 승인 (신뢰도 {confidence_score:.1%}). 경고 {len(warnings)}건 존재하나 진행 가능."
            )
        else:
            return (
                "auto_approve",
                True,
                f"자동 승인 (신뢰도 {confidence_score:.1%}). 문제 없음."
            )

    # Case 4: 신뢰도가 기준 미만이면 needs_review
    return (
        "needs_review",
        False,
        f"수동 검토 필요 (신뢰도 {confidence_score:.1%} < 기준 {threshold:.1%})"
    )

RUN_LOG_PATH = Path(__file__).parent.parent.parent / "training_data" / "windmill_runs.xlsx"


class CallbackPayload(BaseModel):
    status: str
    action: Optional[str] = None
    auto_approve: Optional[bool] = None
    confidence: Optional[float] = None
    message: Optional[str] = None
    file_url: Optional[str] = None
    run_id: Optional[str] = None
    flow_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


class NotificationResult(BaseModel):
    stored: bool
    slack_sent: bool
    email_sent: bool


@router.post("/callback")
async def receive_callback(
    payload: CallbackPayload,
    authorization: Optional[str] = Header(None)
) -> NotificationResult:
    """
    Windmill 플로우 실행 결과를 수신하여 로그로 저장하고 알림을 보냅니다.
    - Excel 로그: training_data/windmill_runs.xlsx (sheet: Sheet1)
    - Slack: SLACK_WEBHOOK_URL 환경변수 설정 시 전송
    - Email: SMTP_HOST/PORT/USER/PASSWORD/EMAIL_TO 모두 설정 시 전송
    
    인증: Authorization: Bearer <WINDMILL_CALLBACK_SECRET>
    """
    # 시크릿 검증 (설정된 경우에만)
    if WINDMILL_CALLBACK_SECRET:
        expected = f"Bearer {WINDMILL_CALLBACK_SECRET}"
        if authorization != expected:
            raise HTTPException(status_code=401, detail="Invalid or missing authorization token")
    
    record = _build_record(payload)

    stored = _persist_run(record)
    slack_sent = _notify_slack(record)
    email_sent = _notify_email(record)

    return NotificationResult(stored=stored, slack_sent=slack_sent, email_sent=email_sent)


@router.get("/latest")
async def latest_runs(limit: int = 10) -> Dict[str, List[Dict[str, Any]]]:
    """최근 실행 기록 조회"""
    return {"runs": _load_runs(limit)}


def _build_record(payload: CallbackPayload) -> Dict[str, Any]:
    return {
        "timestamp": datetime.utcnow().isoformat(),
        "status": payload.status,
        "action": payload.action,
        "auto_approve": payload.auto_approve,
        "confidence": payload.confidence,
        "message": payload.message,
        "file_url": payload.file_url,
        "run_id": payload.run_id,
        "flow_id": payload.flow_id,
        "metadata": payload.metadata or {},
    }


def _persist_run(record: Dict[str, Any]) -> bool:
    """Excel 로그에 한 행 추가 (openpyxl 활용)."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail=f"openpyxl 미설치: {exc}")

    RUN_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "timestamp",
        "status",
        "action",
        "auto_approve",
        "confidence",
        "message",
        "file_url",
        "run_id",
        "flow_id",
        "metadata_json",
    ]

    if RUN_LOG_PATH.exists():
        wb = load_workbook(RUN_LOG_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    ws.append(
        [
            record.get("timestamp"),
            record.get("status"),
            record.get("action"),
            record.get("auto_approve"),
            record.get("confidence"),
            record.get("message"),
            record.get("file_url"),
            record.get("run_id"),
            record.get("flow_id"),
            json.dumps(record.get("metadata", {}), ensure_ascii=False),
        ]
    )

    wb.save(RUN_LOG_PATH)
    return True


def _load_runs(limit: int = 10) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    if not RUN_LOG_PATH.exists():
        return []

    wb = load_workbook(RUN_LOG_PATH)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    records = []
    for row in rows[1:]:
        item = dict(zip(headers, row))
        # metadata_json을 dict로 복원
        metadata_raw = item.get("metadata_json")
        if isinstance(metadata_raw, str):
            try:
                item["metadata"] = json.loads(metadata_raw)
            except json.JSONDecodeError:
                item["metadata"] = metadata_raw
        else:
            item["metadata"] = metadata_raw
        records.append(item)

    return list(reversed(records))[:limit]


def _notify_slack(record: Dict[str, Any]) -> bool:
    webhook_url = os.getenv("SLACK_WEBHOOK_URL")
    if not webhook_url:
        return False

    status_icon = "✅" if record.get("action") == "auto_approve" else "⚠️"
    text = (
        f"{status_icon} Windmill 플로우 결과\n"
        f"- status: {record.get('status')}\n"
        f"- action: {record.get('action')}\n"
        f"- auto_approve: {record.get('auto_approve')}\n"
        f"- confidence: {record.get('confidence')}\n"
        f"- message: {record.get('message') or '-'}\n"
        f"- file: {record.get('file_url') or '-'}\n"
        f"- run_id: {record.get('run_id') or '-'}"
    )

    try:
        resp = requests.post(webhook_url, json={"text": text}, timeout=5)
        return resp.status_code < 300
    except Exception as exc:
        secure_logger.warning(f"Slack 알림 실패: {exc}")
        return False


def _notify_email(record: Dict[str, Any]) -> bool:
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "0") or 0)
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    email_to = os.getenv("EMAIL_TO")

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, email_to]):
        return False

    from email.message import EmailMessage
    import smtplib

    msg = EmailMessage()
    msg["Subject"] = "WIKISOFT3 검증 결과 알림"
    msg["From"] = smtp_user
    msg["To"] = email_to

    body = (
        f"status: {record.get('status')}\n"
        f"action: {record.get('action')}\n"
        f"auto_approve: {record.get('auto_approve')}\n"
        f"confidence: {record.get('confidence')}\n"
        f"message: {record.get('message') or '-'}\n"
        f"file: {record.get('file_url') or '-'}\n"
        f"run_id: {record.get('run_id') or '-'}\n"
        f"flow_id: {record.get('flow_id') or '-'}\n"
        f"timestamp: {record.get('timestamp')}"
    )
    msg.set_content(body)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        return True
    except Exception as exc:
        secure_logger.warning(f"Email 알림 실패: {exc}")
        return False
