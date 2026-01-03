from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
from pydantic import BaseModel
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

router = APIRouter(prefix="/api/export", tags=["export"])


class ExportRequest(BaseModel):
    """수정된 데이터 내보내기 요청"""
    filename: str
    headers: List[str]
    data: List[Dict[str, Any]]


class ExportErrorsRequest(BaseModel):
    """오류 항목만 추출하여 내보내기"""
    filename: str
    errors: List[Dict[str, Any]]  # row, field, message, severity 포함


@router.post("/xlsx")
async def export_xlsx(request: ExportRequest):
    """수정된 데이터를 Excel 파일로 내보내기
    
    Args:
        request: 파일명, 헤더, 데이터
        
    Returns:
        StreamingResponse: Excel 파일 다운로드
    """
    try:
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재직자명부"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col_idx, header in enumerate(request.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, row_data in enumerate(request.data, start=2):
            for col_idx, header in enumerate(request.headers, start=1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(request.headers, start=1):
            max_length = len(str(header))
            for row_idx in range(2, len(request.data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # 최대 50자로 제한
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
        
        # 파일명 생성 (원본명_수정본_날짜.xlsx)
        base_name = request.filename.rsplit(".", 1)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_수정본_{timestamp}.xlsx"
        
        # BytesIO에 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{output_filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 생성 실패: {str(e)}")


@router.post("/errors")
async def export_errors(request: ExportErrorsRequest):
    """검증 오류만 추출하여 Excel 파일로 내보내기
    
    에러와 경고만 포함하여 고객이 수정할 수 있도록 제공
    
    Args:
        request: 파일명, 오류 목록 (row, field, message, severity)
        
    Returns:
        StreamingResponse: 오류 목록 Excel 파일 다운로드
    """
    try:
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "오류항목"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="FF5555", end_color="FF5555", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 헤더 작성: 행번호, 필드명, 오류내용, 심각도
        headers = ["행번호", "필드명", "오류 내용", "심각도"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, error in enumerate(request.errors, start=2):
            # 행번호
            ws.cell(row=row_idx, column=1, value=error.get('row', ''))
            
            # 필드명
            ws.cell(row=row_idx, column=2, value=error.get('field', ''))
            
            # 오류 내용
            message_cell = ws.cell(row=row_idx, column=3, value=error.get('message', ''))
            message_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # 심각도 (색상 구분)
            severity = error.get('severity', 'warning')
            severity_cell = ws.cell(row=row_idx, column=4, value=severity)
            if severity == 'error':
                severity_cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                severity_cell.font = Font(color="CC0000", bold=True)
            else:
                severity_cell.fill = PatternFill(start_color="FFF2D6", end_color="FFF2D6", fill_type="solid")
                severity_cell.font = Font(color="FF8800", bold=True)
            severity_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 10  # 행번호
        ws.column_dimensions['B'].width = 15  # 필드명
        ws.column_dimensions['C'].width = 50  # 오류 내용
        ws.column_dimensions['D'].width = 10  # 심각도
        
        # 행 높이 자동 조정 (텍스트 줄바꿈 대응)
        for row in ws.iter_rows(min_row=2, max_row=len(request.errors) + 1):
            ws.row_dimensions[row[0].row].height = None  # auto
        
        # 파일명 생성 (원본명_의심목록_날짜.xlsx)
        base_name = request.filename.rsplit(".", 1)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base_name}_의심목록_{timestamp}.xlsx"
        
        # BytesIO에 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{output_filename}"
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 생성 실패: {str(e)}")
