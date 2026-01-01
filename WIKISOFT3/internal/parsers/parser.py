from typing import Any, Dict, List, Optional
import csv
import io
import datetime as _dt

import chardet
from openpyxl import load_workbook


# Excel 날짜 직렬 번호 변환 (1900-01-01 기준)
EXCEL_EPOCH = _dt.date(1899, 12, 30)


def _convert_excel_date(value: Any) -> Any:
    """Excel 직렬 번호를 날짜 문자열로 변환."""
    if value is None or value == "":
        return value
    
    # 이미 datetime 객체면 ISO 형식으로 변환
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.isoformat()
    
    # 숫자(직렬번호)면 날짜로 변환 (30000~50000 범위: 1982~2036년)
    try:
        num = float(value)
        if 30000 <= num <= 50000:
            date_val = EXCEL_EPOCH + _dt.timedelta(days=int(num))
            return date_val.isoformat()
    except (ValueError, TypeError):
        pass
    
    # YYYYMMDD 형식 (19670702.0 같은 float)
    try:
        s = str(value).replace(".0", "").strip()
        if len(s) == 8 and s.isdigit():
            year, month, day = int(s[:4]), int(s[4:6]), int(s[6:8])
            if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                return f"{year:04d}-{month:02d}-{day:02d}"
    except (ValueError, TypeError):
        pass
    
    return value


def _normalize_emp_id(value: Any) -> str:
    """사원번호를 정규화: 소수점 제거, 문자열 반환"""
    if value is None or value == "":
        return ""
    s = str(value).strip()
    # 소수점과 그 뒤 숫자 제거 (190001.0 → 190001)
    if '.' in s:
        s = s.split('.')[0]
    return s


def _is_empty_row(row: List[Any], key_columns: List[int] = None) -> bool:
    """행이 비어있는지 확인. key_columns가 지정되면 해당 열만 검사."""
    if key_columns:
        for idx in key_columns:
            if idx < len(row):
                val = row[idx]
                if val is not None and str(val).strip() != "":
                    return False
        return True
    
    # key_columns가 없으면 전체 검사
    for val in row:
        if val is not None and str(val).strip() != "":
            return False
    return True


def _is_description_row(row: List[Any], headers: List[str]) -> bool:
    """설명 행인지 확인 (참고사항에 설명 텍스트가 있고, 사원번호가 비어있는 경우만)."""
    # 참고사항 컬럼 찾기
    ref_idx = -1
    for idx, h in enumerate(headers):
        h_clean = str(h).strip()
        if "참고" in h_clean or "비고" in h_clean:
            ref_idx = idx
            break
    
    if ref_idx < 0 or ref_idx >= len(row):
        return False
    
    ref_val = str(row[ref_idx]).strip()
    if not ref_val or ref_val in ['', 'nan', 'None']:
        return False
    
    # 사원번호 컬럼이 있고 값이 유효하면 데이터 행으로 유지
    emp_idx = -1
    for idx, h in enumerate(headers):
        h_clean = str(h).strip()
        if "사원" in h_clean or "사번" in h_clean:
            emp_idx = idx
            break
    
    if emp_idx >= 0 and emp_idx < len(row):
        emp_val = str(row[emp_idx]).strip()
        if emp_val and emp_val not in ['', 'nan', 'None']:
            # 사원번호가 숫자면 데이터 행
            try:
                float(emp_val.replace(",", ""))
                return False  # 유효한 사원번호 → 데이터 행
            except ValueError:
                pass
    
    # 설명 행의 특징: 참고사항에 특정 키워드 포함
    description_keywords = ["※", "양식", "변경", "입력", "사항"]
    for kw in description_keywords:
        if kw in ref_val:
            return True
    
    return False


def _find_key_columns(headers: List[str]) -> List[int]:
    """핵심 컬럼(사원번호, 성명) 인덱스 찾기."""
    key_keywords = ["사원", "사번", "직원번호", "성명", "이름", "성함"]
    key_indices = []
    for idx, h in enumerate(headers):
        h_clean = str(h).strip().lower()
        for kw in key_keywords:
            if kw in h_clean:
                key_indices.append(idx)
                break
    return key_indices if key_indices else [0]  # 없으면 첫 번째 열


def _normalize_emp_id_in_rows(headers: List[str], rows: List[List[Any]]) -> List[List[Any]]:
    """모든 행의 사원번호를 정규화 (소수점 제거)"""
    # 사원번호 컬럼 찾기
    emp_col_idx = -1
    for idx, h in enumerate(headers):
        h_clean = str(h).strip()
        if "사원" in h_clean or "사번" in h_clean or "직원번호" in h_clean:
            emp_col_idx = idx
            break
    
    if emp_col_idx < 0:
        return rows
    
    # 모든 행의 사원번호 정규화
    normalized_rows = []
    for row in rows:
        new_row = list(row)
        if emp_col_idx < len(new_row):
            new_row[emp_col_idx] = _normalize_emp_id(new_row[emp_col_idx])
        normalized_rows.append(new_row)
    
    return normalized_rows


    """특정 컬럼들의 날짜를 변환."""
    result = list(row)
    for idx in date_columns:
        if idx < len(result):
            result[idx] = _convert_excel_date(result[idx])
    return result


def _find_date_columns(headers: List[str]) -> List[int]:
    """날짜 컬럼 인덱스 찾기."""
    date_keywords = ["일자", "일", "날짜", "생년월일", "입사", "퇴사", "정산일", "지급일", "기준일"]
    date_indices = []
    for idx, h in enumerate(headers):
        h_clean = str(h).strip()
        for kw in date_keywords:
            if kw in h_clean:
                date_indices.append(idx)
                break
    return date_indices

def _find_note_columns(headers: List[str]) -> List[int]:
    """참고사항/비고 컨럼 인덱스 찾기 (제거 대상)."""
    note_keywords = ["참고사항", "참고", "비고", "메모", "비고란", "노트"]
    note_indices = []
    for idx, h in enumerate(headers):
        h_clean = str(h).strip().lower()
        for kw in note_keywords:
            if kw in h_clean:
                note_indices.append(idx)
                break
    return note_indices


def _remove_columns(headers: List[str], rows: List[List[Any]], columns_to_remove: List[int]) -> tuple:
    """특정 컨럼을 헤더와 모든 행에서 제거."""
    if not columns_to_remove:
        return headers, rows
    
    columns_to_remove_set = set(columns_to_remove)
    new_headers = [h for idx, h in enumerate(headers) if idx not in columns_to_remove_set]
    new_rows = []
    for row in rows:
        new_row = [val for idx, val in enumerate(row) if idx not in columns_to_remove_set]
        new_rows.append(new_row)
    
    return new_headers, new_rows

def _infer_types(rows: List[List[Any]], sample_rows: int = 200) -> Dict[int, str]:
    """간단한 컬럼 타입 추론(문자/숫자/날짜 후보)."""
    import datetime as _dt

    col_types: Dict[int, str] = {}
    sample = rows[:sample_rows]
    for col_idx in range(max((len(r) for r in sample), default=0)):
        values = [r[col_idx] for r in sample if col_idx < len(r)]
        num_cnt = 0
        date_cnt = 0
        for v in values:
            if isinstance(v, (int, float)):
                num_cnt += 1
            elif isinstance(v, _dt.date):
                date_cnt += 1
            else:
                # 숫자 문자열
                try:
                    float(str(v).replace(",", ""))
                    num_cnt += 1
                except Exception:  # noqa: BLE001
                    pass
        if date_cnt > max(num_cnt, 0):
            col_types[col_idx] = "date"
        elif num_cnt > 0:
            col_types[col_idx] = "number"
        else:
            col_types[col_idx] = "string"
    return col_types


def _parse_csv(text: str) -> Dict[str, Any]:
    reader = csv.reader(io.StringIO(text))
    all_rows: List[List[str]] = list(reader)
    headers = all_rows[0] if all_rows else []
    raw_data_rows = all_rows[1:] if len(all_rows) > 1 else []
    
    # 날짜/키 컬럼 찾기
    key_columns = _find_key_columns(headers)
    date_columns = _find_date_columns(headers)
    
    # 빈 행 필터링 + 날짜 변환
    data_rows = []
    skipped_empty = 0
    
    for raw_row in raw_data_rows:
        if _is_empty_row(raw_row, key_columns):
            skipped_empty += 1
            continue
        converted_row = _convert_dates_in_row(raw_row, date_columns)
        data_rows.append(converted_row)
    
    # 사원번호 정규화
    data_rows = _normalize_emp_id_in_rows(headers, data_rows)
    
    return {
        "headers": headers,
        "rows": data_rows,
        "meta": {
            "parser": "csv",
            "total_rows": len(data_rows),
            "skipped_empty_rows": skipped_empty,
            "column_types": _infer_types(data_rows),
        },
    }


def _parse_xlsx(file_bytes: bytes, sheet_name: Optional[str] = None, max_rows: int = 5000) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows: List[List[Any]] = []
    headers: List[Any] = []
    key_columns: List[int] = []
    date_columns: List[int] = []
    skipped_empty = 0
    skipped_desc = 0
    
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            # 헤더 정리: 줄바꿈/공백 제거
            headers = [
                "" if c is None else str(c).replace('\n', ' ').replace('\r', '').strip()
                for c in row
            ]
            key_columns = _find_key_columns(headers)
            date_columns = _find_date_columns(headers)
            continue
        if max_rows and len(rows) >= max_rows:
            break
        
        raw_row = ["" if c is None else c for c in row]
        
        # 빈 행 필터링
        if _is_empty_row(raw_row, key_columns):
            skipped_empty += 1
            continue
        
        # 설명 행 필터링
        if _is_description_row(raw_row, headers):
            skipped_desc += 1
            continue
        
        # 날짜 컬럼 변환
        converted_row = _convert_dates_in_row(raw_row, date_columns)
        rows.append(converted_row)
        # 참고사항/비고 열 제거
    note_columns = _find_note_columns(headers)
    if note_columns:
        headers, rows = _remove_columns(headers, rows, note_columns)
    
    # 사원번호 정규화
    rows = _normalize_emp_id_in_rows(headers, rows)
    
    return {
        "headers": headers,
        "rows": rows,
        "meta": {
            "parser": "xlsx",
            "total_rows_sampled": len(rows),
            "skipped_empty_rows": skipped_empty,
            "skipped_description_rows": skipped_desc,
            "sheet": ws.title,
            "column_types": _infer_types(rows),
            "note": f"capped at {max_rows} rows, {skipped_empty} empty + {skipped_desc} desc rows filtered",
        },
    }


def _parse_xls(file_bytes: bytes, sheet_name: Optional[str] = None, max_rows: int = 5000) -> Dict[str, Any]:
    """XLS (구버전 Excel) 파싱 - pandas + xlrd 사용."""
    import pandas as pd

    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd")
    target_sheet = sheet_name if sheet_name and sheet_name in xls.sheet_names else xls.sheet_names[0]

    # 재직자 명부 시트 자동 탐색 - 우선순위: (2-2) > 재직자명부_시스템input > 재직자+명부
    priority_patterns = [
        ("(2-2)", "재직자"),  # 가장 정확한 패턴
        ("재직자명부", "시스템"),  # 시스템 input
    ]
    
    for patterns in priority_patterns:
        for name in xls.sheet_names:
            if all(p in name for p in patterns):
                target_sheet = name
                break
        else:
            continue
        break
    else:
        # fallback: 재직자+명부
        for name in xls.sheet_names:
            if "재직자" in name and "명부" in name:
                target_sheet = name
            break

    df = pd.read_excel(xls, sheet_name=target_sheet, header=0, nrows=max_rows)
    # 헤더 정리: 줄바꿈/공백 제거
    headers = [str(c).replace('\n', ' ').replace('\r', '').strip() for c in df.columns.tolist()]
    
    # 날짜/키 컬럼 찾기
    key_columns = _find_key_columns(headers)
    date_columns = _find_date_columns(headers)
    
    # 빈 행 필터링 + 날짜 변환
    raw_rows = df.values.tolist()
    rows = []
    skipped_empty = 0
    skipped_desc = 0
    
    for raw_row in raw_rows:
        # NaN을 빈 문자열로 변환
        clean_row = ["" if (pd.isna(v) if hasattr(pd, 'isna') else v != v) else v for v in raw_row]
        
        if _is_empty_row(clean_row, key_columns):
            skipped_empty += 1
            continue
        
        # 설명 행 필터링
        if _is_description_row(clean_row, headers):
            skipped_desc += 1
            continue
        
        # 날짜 컬럼 변환
        converted_row = _convert_dates_in_row(clean_row, date_columns)
        rows.append(converted_row)

    # 참고사항/비고 열 제거
    note_columns = _find_note_columns(headers)
    if note_columns:
        headers, rows = _remove_columns(headers, rows, note_columns)

    # 사원번호 정규화
    rows = _normalize_emp_id_in_rows(headers, rows)

    return {
        "headers": headers,
        "rows": rows,
        "meta": {
            "parser": "xls",
            "total_rows_sampled": len(rows),
            "skipped_empty_rows": skipped_empty,
            "skipped_description_rows": skipped_desc,
            "sheet": target_sheet,
            "available_sheets": xls.sheet_names,
            "column_types": _infer_types(rows),
            "note": f"capped at {max_rows} rows, {skipped_empty} empty + {skipped_desc} desc rows filtered",
        },
    }


def parse_roster(file_bytes: bytes, sheet_name: Optional[str] = None, max_rows: int = 5000) -> Dict[str, Any]:
    """CSV/xlsx/xls 파서 (스트리밍 샘플 기반).

    - xlsx: read_only 모드로 최대 max_rows 샘플링, 시트 선택 지원.
    - xls: pandas + xlrd로 구버전 Excel 지원.
    - csv: chardet로 인코딩 감지 후 파싱.
    """
    # xlsx는 ZIP(0x50 0x4B) 시그니처
    if file_bytes[:2] == b"PK":
        return _parse_xlsx(file_bytes, sheet_name=sheet_name, max_rows=max_rows)

    # xls는 OLE2 (0xD0 0xCF) 시그니처
    if file_bytes[:2] == b"\xd0\xcf":
        return _parse_xls(file_bytes, sheet_name=sheet_name, max_rows=max_rows)

    # 나머지는 CSV로 시도
    detected = chardet.detect(file_bytes)
    encoding = detected.get("encoding") or "utf-8"
    text = file_bytes.decode(encoding, errors="replace")
    parsed = _parse_csv(text)
    parsed["meta"]["encoding"] = encoding
    return parsed
