"""
Excel 리포트 생성기

검증 결과를 기반으로 Excel 파일 생성:
- 빨간색: 오류 (필수값 누락, 데이터 불일치)
- 노란색: 경고 (5% 이상 차이, 이상치 탐지)
- 초록색: 정상
"""

from typing import Any, Dict, List, Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json


# 셀 스타일 정의
STYLE_ERROR = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 빨간색
STYLE_WARNING = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")  # 노란색
STYLE_SUCCESS = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")  # 초록색
STYLE_HEADER = PatternFill(start_color="4C6EF5", end_color="4C6EF5", fill_type="solid")  # 파란색

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(size=10)
FONT_ERROR = Font(color="8B0000", bold=True)
FONT_WARNING = Font(color="8B4513")

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')


def generate_report(validation: Dict[str, Any]) -> Dict[str, Any]:
    """JSON 리포트 생성"""
    return {
        "report_type": "json",
        "summary": {
            "passed": validation.get("passed"),
            "warnings": validation.get("warnings", []),
            "checks": validation.get("checks", []),
        },
    }


def generate_excel_report(
    validation_result: Dict[str, Any],
    original_data: Optional[Dict[str, Any]] = None,
    answers: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    검증 결과를 기반으로 Excel 리포트 생성
    
    Args:
        validation_result: 검증 결과 (steps, confidence, anomalies 등)
        original_data: 원본 파싱 데이터 (headers, rows)
        answers: 사용자 진단 답변
    
    Returns:
        Excel 파일 bytes
    """
    wb = Workbook()
    
    # 1. 요약 시트
    ws_summary = wb.active
    ws_summary.title = "검증 요약"
    _create_summary_sheet(ws_summary, validation_result, answers)
    
    # 2. 매칭 결과 시트
    ws_matching = wb.create_sheet("헤더 매칭")
    _create_matching_sheet(ws_matching, validation_result)
    
    # 3. 이상 탐지 시트
    if validation_result.get("anomalies", {}).get("detected"):
        ws_anomalies = wb.create_sheet("이상 탐지")
        _create_anomalies_sheet(ws_anomalies, validation_result)
    
    # 4. 원본 데이터 시트 (하이라이팅 포함)
    if original_data and original_data.get("rows"):
        ws_data = wb.create_sheet("검증된 데이터")
        _create_data_sheet(ws_data, original_data, validation_result)
    
    # 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_summary_sheet(ws, validation_result: Dict, answers: Optional[Dict]):
    """요약 시트 생성"""
    # 제목
    ws.merge_cells('A1:D1')
    ws['A1'] = "🏢 WIKISOFT3 검증 결과 리포트"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = ALIGN_CENTER
    
    # 검증 상태
    row = 3
    status = validation_result.get("status", "unknown")
    confidence = validation_result.get("confidence", {})
    
    headers = ["항목", "값", "상태", "설명"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER
    
    # 데이터 행
    summary_data = [
        ("검증 상태", status, "✅" if status == "ok" else "❌", "전체 검증 결과"),
        ("신뢰도 점수", f"{confidence.get('score', 0) * 100:.1f}%", 
         "✅" if confidence.get('score', 0) >= 0.8 else "⚠️", confidence.get('grade', '')),
        ("분석 행 수", validation_result.get("steps", {}).get("parsed_summary", {}).get("row_count", 0),
         "✅", "파싱된 데이터 행 수"),
        ("이상 탐지", len(validation_result.get("anomalies", {}).get("anomalies", [])),
         "⚠️" if validation_result.get("anomalies", {}).get("detected") else "✅", 
         validation_result.get("anomalies", {}).get("recommendation", "")),
    ]
    
    for data_row in summary_data:
        row += 1
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_LEFT
            
            # 상태 컬럼 스타일링
            if col == 3:
                if "❌" in str(value):
                    cell.fill = STYLE_ERROR
                elif "⚠️" in str(value):
                    cell.fill = STYLE_WARNING
                else:
                    cell.fill = STYLE_SUCCESS
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    
    # 진단 답변 섹션
    if answers:
        row += 3
        ws.cell(row=row, column=1, value="📋 진단 답변").font = Font(bold=True, size=12)
        row += 1
        
        for q_id, answer in answers.items():
            row += 1
            ws.cell(row=row, column=1, value=q_id)
            ws.cell(row=row, column=2, value=str(answer))


def _create_matching_sheet(ws, validation_result: Dict):
    """헤더 매칭 결과 시트"""
    matches = validation_result.get("steps", {}).get("matches", {}).get("matches", [])
    
    # 헤더
    headers = ["원본 헤더", "매칭된 필드", "신뢰도", "AI 사용", "상태"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
    
    # 데이터
    for row_idx, match in enumerate(matches, 2):
        source = match.get("source", "")
        target = match.get("target", "")
        confidence = match.get("confidence", 0)
        used_ai = match.get("used_ai", False)
        unmapped = match.get("unmapped", False)
        
        ws.cell(row=row_idx, column=1, value=source).border = BORDER_THIN
        ws.cell(row=row_idx, column=2, value=target or "(미매칭)").border = BORDER_THIN
        
        conf_cell = ws.cell(row=row_idx, column=3, value=f"{confidence * 100:.0f}%")
        conf_cell.border = BORDER_THIN
        
        ws.cell(row=row_idx, column=4, value="예" if used_ai else "아니오").border = BORDER_THIN
        
        status_cell = ws.cell(row=row_idx, column=5)
        status_cell.border = BORDER_THIN
        
        if unmapped or not target:
            status_cell.value = "❌ 미매칭"
            status_cell.fill = STYLE_ERROR
        elif confidence < 0.7:
            status_cell.value = "⚠️ 낮은 신뢰도"
            status_cell.fill = STYLE_WARNING
        else:
            status_cell.value = "✅ 매칭"
            status_cell.fill = STYLE_SUCCESS
    
    # 컬럼 너비
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15


def _create_anomalies_sheet(ws, validation_result: Dict):
    """이상 탐지 시트"""
    anomalies = validation_result.get("anomalies", {}).get("anomalies", [])
    
    # 헤더
    headers = ["유형", "심각도", "메시지", "필드", "값"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
    
    # 데이터
    for row_idx, anomaly in enumerate(anomalies, 2):
        ws.cell(row=row_idx, column=1, value=anomaly.get("type", "")).border = BORDER_THIN
        
        severity_cell = ws.cell(row=row_idx, column=2, value=anomaly.get("severity", ""))
        severity_cell.border = BORDER_THIN
        if anomaly.get("severity") == "high":
            severity_cell.fill = STYLE_ERROR
        elif anomaly.get("severity") == "medium":
            severity_cell.fill = STYLE_WARNING
        
        ws.cell(row=row_idx, column=3, value=anomaly.get("message", "")).border = BORDER_THIN
        ws.cell(row=row_idx, column=4, value=anomaly.get("field", "")).border = BORDER_THIN
        ws.cell(row=row_idx, column=5, value=str(anomaly.get("value", ""))).border = BORDER_THIN
    
    # 컬럼 너비
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20


def _create_data_sheet(ws, original_data: Dict, validation_result: Dict):
    """
    원본 데이터 시트 (하이라이팅 포함)
    
    빨간색: 오류 셀 (필수값 누락, 형식 오류)
    노란색: 경고 셀 (이상치, 5% 초과 차이)
    """
    headers = original_data.get("headers", [])
    rows = original_data.get("rows", [])
    
    # 이상 탐지 정보로 하이라이팅할 셀 결정
    anomalies = validation_result.get("anomalies", {}).get("anomalies", [])
    error_cells = set()  # (row, col) 튜플
    warning_cells = set()
    
    for anomaly in anomalies:
        field = anomaly.get("field", "")
        row_idx = anomaly.get("row", None)
        severity = anomaly.get("severity", "medium")
        
        if field in headers:
            col_idx = headers.index(field)
            if row_idx is not None:
                if severity == "high":
                    error_cells.add((row_idx + 2, col_idx + 1))  # +2: 헤더 행 오프셋
                else:
                    warning_cells.add((row_idx + 2, col_idx + 1))
    
    # 헤더 작성
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = STYLE_HEADER
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) + 2)
    
    # 데이터 작성
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_THIN
            
            # 하이라이팅
            if (row_idx, col_idx) in error_cells:
                cell.fill = STYLE_ERROR
                cell.font = FONT_ERROR
            elif (row_idx, col_idx) in warning_cells:
                cell.fill = STYLE_WARNING
                cell.font = FONT_WARNING
            
            # 빈 필수값 체크
            if value is None or (isinstance(value, str) and value.strip() == ""):
                # 필수 필드인 경우 빨간색
                header = headers[col_idx - 1] if col_idx <= len(headers) else ""
                required_fields = ["사원번호", "이름", "생년월일", "입사일자", "기준급여"]
                if header in required_fields:
                    cell.fill = STYLE_ERROR
                    cell.value = "(누락)"


def export_validation_to_excel(
    validation_result: Dict[str, Any],
    original_data: Optional[Dict[str, Any]] = None,
    answers: Optional[Dict[str, Any]] = None,
    output_path: Optional[str] = None
) -> bytes:
    """
    검증 결과를 Excel 파일로 내보내기
    
    Args:
        validation_result: 검증 결과
        original_data: 원본 파싱 데이터
        answers: 진단 답변
        output_path: 저장 경로 (없으면 bytes 반환)
    
    Returns:
        Excel 파일 bytes
    """
    excel_bytes = generate_excel_report(validation_result, original_data, answers)
    
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(excel_bytes)
    
    return excel_bytes
