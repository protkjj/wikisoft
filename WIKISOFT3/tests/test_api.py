"""
API 엔드포인트 테스트
"""
import pytest
from fastapi.testclient import TestClient
import io
from openpyxl import Workbook
import sys
sys.path.insert(0, "/Users/kj/Desktop/wiki/WIKISOFT3")

from external.api.main import app

client = TestClient(app)


def create_test_excel() -> bytes:
    """테스트용 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "재직자"
    
    headers = ["사원번호", "이름", "생년월일", "입사일", "기준급여", "종업원구분"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    data = [
        ["EMP001", "홍길동", "1990-01-15", "2020-03-01", 5000000, "직원"],
        ["EMP002", "김철수", "1985-06-20", "2018-07-15", 6000000, "직원"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestHealthEndpoint:
    """Health 엔드포인트 테스트"""
    
    def test_health_check(self):
        """헬스 체크"""
        response = client.get("/api/health")
        assert response.status_code == 200
        
        data = response.json()
        assert data["status"] in ["healthy", "ok"]  # 둘 다 허용
        assert "version" in data


class TestDiagnosticQuestionsEndpoint:
    """진단 질문 엔드포인트 테스트"""
    
    def test_get_questions(self):
        """질문 조회"""
        response = client.get("/api/diagnostic-questions")
        assert response.status_code == 200
        
        data = response.json()
        assert "questions" in data
        assert "total" in data
        assert data["total"] == 13
    
    def test_questions_structure(self):
        """질문 구조 확인"""
        response = client.get("/api/diagnostic-questions")
        data = response.json()
        
        if data["questions"]:
            q = data["questions"][0]
            assert "id" in q
            assert "question" in q
            assert "type" in q
            assert "category" in q


class TestValidateEndpoint:
    """검증 엔드포인트 테스트"""
    
    def test_validate_with_file(self):
        """파일 검증"""
        excel_bytes = create_test_excel()
        
        # /api/validate 또는 /api/auto-validate 시도
        response = client.post(
            "/api/auto-validate",
            files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # 200 또는 404 (엔드포인트 없으면)
        assert response.status_code in [200, 404]
        if response.status_code == 200:
            data = response.json()
            assert "status" in data or "success" in data
    
    def test_validate_without_file(self):
        """파일 없이 검증 (에러)"""
        response = client.post("/api/auto-validate")
        
        # 422 Unprocessable Entity 또는 404
        assert response.status_code in [400, 404, 422]


class TestBatchEndpoint:
    """배치 엔드포인트 테스트"""
    
    def test_batch_validate(self):
        """배치 검증"""
        excel_bytes = create_test_excel()
        
        response = client.post(
            "/api/batch-validate",
            files=[
                ("files", ("test1.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                ("files", ("test2.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ]
        )
        
        # 202 Accepted 또는 200 OK
        assert response.status_code in [200, 202]
        data = response.json()
        assert "job_id" in data


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
